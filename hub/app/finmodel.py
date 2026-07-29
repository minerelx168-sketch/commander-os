"""CFO scenario model — turns a board debate into a live Excel forecast.

Division of labour (deliberate):

* the **LLM supplies assumptions only** — revenue drivers, cost structure,
  capex, and the three scenario multipliers, as JSON. It never does the
  arithmetic; language models get compound growth and break-even wrong, and a
  number the CEO cannot trace is worse than no number.
* **Python writes real Excel formulas** — every derived cell is an `=` formula
  referencing the Assumptions sheet, so the CEO changes one input and the whole
  model, all three scenarios, and the charts move with it. A workbook of frozen
  literals would be a report; this is a model you can interrogate.

Sheets produced:
  1. สมมติฐาน (Assumptions)  — the only place with typed-in numbers, all named
  2. Base / Best / Worst      — 24-month P&L + cash flow, formula-driven
  3. เปรียบเทียบ Scenario     — side-by-side with charts
  4. ความเสี่ยง (Sensitivity) — 2-way data table + break-even + runway
  5. วิธีอ่าน (How to read)   — provenance, guardrails, what to change
"""
import io
import json
import logging
import re

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config, llm, store

log = logging.getLogger("hub.finmodel")

MONTHS = 24

# ── Jarvis-ish palette, readable when printed ──
_NAVY = "FF0B1F33"
_CYAN = "FF00A6C8"
_GREY = "FFF2F5F7"
_WARN = "FFFFF4E5"
_BAD = "FFFDE7E9"
_GOOD = "FFE8F6EE"
_THIN = Side(style="thin", color="FFD0D7DE")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BAHT = '#,##0;[Red]-#,##0'
_PCT = '0.0%'

ASSUMPTION_SYSTEM = (
    "คุณคือ CFO ที่สร้างแบบจำลองการเงินให้ CEO ใช้ตัดสินใจจริง\n"
    "หน้าที่ของคุณคือสกัด 'สมมติฐานเชิงตัวเลข' จากบทถกเถียงของบอร์ด ให้กลายเป็น input "
    "ของแบบจำลอง 24 เดือน — คุณไม่ต้องคำนวณอะไรเลย ระบบจะคำนวณเองด้วยสูตร Excel\n"
    "\n"
    "กฎเหล็กเรื่องความซื่อสัตย์ของตัวเลข (สำคัญที่สุด):\n"
    "- ตัวเลขใดที่ 'ปรากฏในบทถกเถียงหรือเอกสาร' ให้ใส่ source เป็นข้อความอ้างอิงสั้นๆ ว่ามาจากไหน\n"
    "- ตัวเลขใดที่คุณต้องประมาณเอง ให้ระบุ source ว่า \"ประมาณการของ CFO\" ตรงๆ ห้ามอ้างว่ามาจากบอร์ด\n"
    "- ห้ามแต่งตัวเลขให้ดูน่าเชื่อ ถ้าไม่มีข้อมูลจริงให้ใส่ค่าที่อนุรักษ์นิยมและบอกว่าเป็นการประมาณ\n"
    "- ทุกค่าเป็นหน่วยบาท (THB) ยกเว้นที่ระบุว่าเป็น % หรือจำนวนหน่วย\n"
    "\n"
    "ตอบเป็น JSON ล้วนเท่านั้น ห้ามมีข้อความอื่นนอก JSON:\n"
    "{\n"
    '  "business_model": "ธุรกิจนี้หาเงินจากอะไร 1 ประโยค",\n'
    '  "currency_note": "หมายเหตุเรื่องหน่วยเงินและช่วงเวลา",\n'
    '  "revenue": {\n'
    '    "units_month_1": {"value": 100, "source": "...", "note": "จำนวนหน่วยที่ขายได้เดือนแรก"},\n'
    '    "price_per_unit": {"value": 100, "source": "...", "note": "ราคาขายต่อหน่วย (บาท)"},\n'
    '    "growth_rate_monthly": {"value": 0.03, "source": "...", "note": "อัตราโตต่อเดือน เช่น 0.03 = 3%"},\n'
    '    "churn_or_seasonality": {"value": 0.0, "source": "...", "note": "อัตราหดตัว/ฤดูกาลต่อเดือน"}\n'
    "  },\n"
    '  "costs": {\n'
    '    "cogs_pct_of_revenue": {"value": 0.4, "source": "...", "note": "ต้นทุนขายเป็น % ของรายได้"},\n'
    '    "fixed_cost_month": {"value": 20000, "source": "...", "note": "ค่าใช้จ่ายคงที่ต่อเดือน (เช่า+เงินเดือน)"},\n'
    '    "marketing_pct_of_revenue": {"value": 0.1, "source": "...", "note": "งบตลาดเป็น % ของรายได้"},\n'
    '    "other_variable_pct": {"value": 0.05, "source": "...", "note": "ต้นทุนผันแปรอื่น % ของรายได้"}\n'
    "  },\n"
    '  "capital": {\n'
    '    "upfront_investment": {"value": 0, "source": "...", "note": "เงินลงทุนเริ่มต้น (บาท)"},\n'
    '    "starting_cash": {"value": 0, "source": "...", "note": "เงินสดตั้งต้น (บาท)"},\n'
    '    "payment_collection_lag_months": {"value": 0, "source": "...", "note": "เก็บเงินช้ากี่เดือน"}\n'
    "  },\n"
    '  "scenarios": {\n'
    '    "base": {"revenue_mult": 1.0, "cost_mult": 1.0, "rationale": "กรณีฐาน"},\n'
    '    "best": {"revenue_mult": 1.25, "cost_mult": 0.95, "rationale": "อะไรต้องเกิดจริงถึงจะได้กรณีนี้"},\n'
    '    "worst": {"revenue_mult": 0.7, "cost_mult": 1.15, "rationale": "อะไรที่ทำให้แย่ลงได้จริง"}\n'
    "  },\n"
    '  "risks": [{"risk": "ความเสี่ยงที่วัดเป็นตัวเลขได้",\n'
    '             "driver": "ตัวแปรไหนในแบบจำลองที่สะท้อนความเสี่ยงนี้",\n'
    '             "trigger": "สัญญาณเตือนที่ CEO ต้องเฝ้า (เชิงตัวเลข)",\n'
    '             "mitigation": "ทำอะไรถ้าเกิดขึ้น"}],\n'
    '  "kpis_to_watch": ["ตัวเลขที่ CEO ต้องดูทุกเดือน"],\n'
    '  "sensitivity": {"driver_a": "revenue.growth_rate_monthly",\n'
    '                  "driver_b": "costs.cogs_pct_of_revenue",\n'
    '                  "why": "ทำไมสองตัวนี้คือตัวแปรที่ชี้เป็นชี้ตาย"}\n'
    "}"
)

# Assumption rows: (json_path, label, excel_number_format, defined_name)
_ROWS = [
    ("revenue.units_month_1", "จำนวนหน่วยที่ขายเดือนแรก", "#,##0", "units_m1"),
    ("revenue.price_per_unit", "ราคาขายต่อหน่วย (บาท)", _BAHT, "price"),
    ("revenue.growth_rate_monthly", "อัตราเติบโตต่อเดือน", _PCT, "growth"),
    ("revenue.churn_or_seasonality", "อัตราหดตัว/ฤดูกาลต่อเดือน", _PCT, "churn"),
    ("costs.cogs_pct_of_revenue", "ต้นทุนขาย (% ของรายได้)", _PCT, "cogs_pct"),
    ("costs.fixed_cost_month", "ค่าใช้จ่ายคงที่ต่อเดือน (บาท)", _BAHT, "fixed_cost"),
    ("costs.marketing_pct_of_revenue", "งบการตลาด (% ของรายได้)", _PCT, "mkt_pct"),
    ("costs.other_variable_pct", "ต้นทุนผันแปรอื่น (% ของรายได้)", _PCT, "other_pct"),
    ("capital.upfront_investment", "เงินลงทุนเริ่มต้น (บาท)", _BAHT, "capex"),
    ("capital.starting_cash", "เงินสดตั้งต้น (บาท)", _BAHT, "cash0"),
    ("capital.payment_collection_lag_months", "ระยะเวลาเก็บเงิน (เดือน)", "0", "collect_lag"),
]

_SCENARIOS = [("base", "Base", _GREY), ("best", "Best", _GOOD), ("worst", "Worst", _BAD)]


def _dig(data: dict, path: str):
    cur = data
    for part in path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur


def _num(entry, default=0.0) -> float:
    """Assumption cells arrive as {"value": n, ...} but tolerate a bare number."""
    if isinstance(entry, dict):
        entry = entry.get("value")
    if isinstance(entry, bool) or entry is None:
        return default
    if isinstance(entry, (int, float)):
        return float(entry)
    try:  # "1,250 บาท" / "3%" — strip anything that isn't numeric
        txt = str(entry).strip()
        pct = txt.endswith("%")
        val = float(re.sub(r"[^0-9.\-]", "", txt) or 0)
        return val / 100 if pct else val
    except ValueError:
        return default


def _src(entry) -> str:
    if isinstance(entry, dict):
        return str(entry.get("source") or entry.get("note") or "—")
    return "—"


def _note(entry) -> str:
    return str(entry.get("note") or "") if isinstance(entry, dict) else ""


# ── sheet builders ──

def _title(ws, text: str, span: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14, color="FFFFFFFF")
    c.fill = PatternFill("solid", fgColor=_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _hdr(ws, row: int, values: list, fill: str = _CYAN) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=True, color="FFFFFFFF" if fill == _CYAN else "FF000000", size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BOX


def _assumptions_sheet(wb: Workbook, model: dict) -> dict:
    """The single source of truth. Returns {defined_name: 'สมมติฐาน!$B$n'}."""
    ws = wb.create_sheet("สมมติฐาน")
    _title(ws, "สมมติฐานของแบบจำลอง — แก้ได้เฉพาะช่องสีเหลืองในคอลัมน์ B", 5)
    ws["A3"] = "แก้ตัวเลขในคอลัมน์ B แล้วทุกชีตและกราฟจะคำนวณใหม่ทั้งหมดอัตโนมัติ"
    ws["A3"].font = Font(italic=True, size=9, color="FF57606A")
    ws["A4"] = f"ธุรกิจ: {model.get('business_model', '—')}"
    ws["A4"].font = Font(size=9, color="FF57606A")

    _hdr(ws, 6, ["ตัวแปร", "ค่า", "ที่มาของตัวเลข", "หมายเหตุ", "ประเภท"])
    refs, row = {}, 7
    for path, label, fmt, name in _ROWS:
        entry = _dig(model, path)
        ws.cell(row=row, column=1, value=label).font = Font(size=10)
        cell = ws.cell(row=row, column=2, value=_num(entry))
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor="FFFFF9DB")  # editable = yellow
        cell.font = Font(bold=True, size=10)
        src = _src(entry)
        ws.cell(row=row, column=3, value=src).font = Font(size=9)
        ws.cell(row=row, column=4, value=_note(entry)).font = Font(size=9, color="FF57606A")
        est = "ประมาณการ" if "ประมาณ" in src or src == "—" else "จากบทถกเถียง"
        c5 = ws.cell(row=row, column=5, value=est)
        c5.font = Font(size=9, bold=est == "ประมาณการ",
                       color="FFB35C00" if est == "ประมาณการ" else "FF1A7F37")
        if est == "ประมาณการ":
            c5.fill = PatternFill("solid", fgColor=_WARN)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = _BOX
        refs[name] = f"'สมมติฐาน'!$B${row}"
        row += 1

    # scenario multipliers live here too so the CEO can stress them directly
    row += 1
    _hdr(ws, row, ["Scenario", "ตัวคูณรายได้", "ตัวคูณต้นทุน", "เหตุผล / เงื่อนไขที่ต้องเป็นจริง", ""])
    row += 1
    for key, label, fill in _SCENARIOS:
        sc = (model.get("scenarios") or {}).get(key) or {}
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=fill)
        for col, field, dflt in ((2, "revenue_mult", 1.0), (3, "cost_mult", 1.0)):
            c = ws.cell(row=row, column=col, value=_num(sc.get(field), dflt))
            c.number_format = "0.00"
            c.fill = PatternFill("solid", fgColor="FFFFF9DB")
            c.font = Font(bold=True, size=10)
        ws.cell(row=row, column=4, value=str(sc.get("rationale") or "—")).font = Font(size=9)
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = _BOX
        refs[f"{key}_rev_mult"] = f"'สมมติฐาน'!$B${row}"
        refs[f"{key}_cost_mult"] = f"'สมมติฐาน'!$C${row}"
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 15
    ws.freeze_panes = "A7"
    return refs


def _scenario_sheet(wb: Workbook, name: str, label: str, refs: dict, fill: str) -> dict:
    """24-month P&L + cash flow built entirely from formulas."""
    ws = wb.create_sheet(label)
    _title(ws, f"แบบจำลอง 24 เดือน — Scenario {label}", MONTHS + 2)
    ws["A3"] = "ทุกช่องในชีตนี้เป็นสูตร ไม่มีเลขพิมพ์มือ — แก้ที่ชีต 'สมมติฐาน' เท่านั้น"
    ws["A3"].font = Font(italic=True, size=9, color="FF57606A")

    rev_m, cost_m = refs[f"{name}_rev_mult"], refs[f"{name}_cost_mult"]
    _hdr(ws, 5, ["รายการ"] + [f"เดือน {i}" for i in range(1, MONTHS + 1)] + ["รวม 24 เดือน"], fill)

    # row layout: label -> formula template keyed on the column letter
    def r(row: int, label: str, tmpl, fmt=_BAHT, bold=False, top=False) -> int:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(bold=bold, size=10)
        c.border = _BOX
        for i in range(MONTHS):
            col = get_column_letter(2 + i)
            cell = ws.cell(row=row, column=2 + i, value=tmpl(col, i))
            cell.number_format = fmt
            cell.font = Font(bold=bold, size=10)
            cell.border = _BOX
            if top:
                cell.fill = PatternFill("solid", fgColor=fill)
        last = get_column_letter(MONTHS + 2)
        tot = ws.cell(row=row, column=MONTHS + 2,
                      value=f"=SUM(B{row}:{get_column_letter(MONTHS + 1)}{row})")
        tot.number_format = fmt
        tot.font = Font(bold=True, size=10)
        tot.fill = PatternFill("solid", fgColor=_GREY)
        tot.border = _BOX
        ws.column_dimensions[last].width = 16
        return row + 1

    row = 6
    # units compound growth net of churn, scaled by the scenario's revenue multiplier
    units = row
    row = r(row, "จำนวนหน่วยที่ขาย", lambda col, i: (
        f"={refs['units_m1']}*{rev_m}" if i == 0
        else f"={get_column_letter(1 + i)}{units}*(1+{refs['growth']}-{refs['churn']})"
    ), "#,##0")
    rev = row
    row = r(row, "รายได้", lambda col, i: f"={col}{units}*{refs['price']}", _BAHT, True, True)
    cogs = row
    row = r(row, "ต้นทุนขาย (COGS)", lambda col, i: f"=-{col}{rev}*{refs['cogs_pct']}*{cost_m}")
    gp = row
    row = r(row, "กำไรขั้นต้น", lambda col, i: f"={col}{rev}+{col}{cogs}", _BAHT, True)
    mkt = row
    row = r(row, "ค่าการตลาด", lambda col, i: f"=-{col}{rev}*{refs['mkt_pct']}*{cost_m}")
    oth = row
    row = r(row, "ต้นทุนผันแปรอื่น", lambda col, i: f"=-{col}{rev}*{refs['other_pct']}*{cost_m}")
    fx = row
    row = r(row, "ค่าใช้จ่ายคงที่", lambda col, i: f"=-{refs['fixed_cost']}*{cost_m}")
    ni = row
    row = r(row, "กำไรสุทธิ (Net Profit)", lambda col, i:
            f"={col}{gp}+{col}{mkt}+{col}{oth}+{col}{fx}", _BAHT, True, True)
    margin = row
    row = r(row, "อัตรากำไรสุทธิ", lambda col, i:
            f"=IFERROR({col}{ni}/{col}{rev},0)", _PCT)

    row += 1
    _hdr(ws, row, ["กระแสเงินสด"] + [f"เดือน {i}" for i in range(1, MONTHS + 1)] + ["—"], fill)
    row += 1
    # collection lag shifts cash in, so cash flow != profit
    cin = row
    row = r(row, "เงินสดรับ (หลังหักระยะเก็บเงิน)", lambda col, i: (
        f"=IF({refs['collect_lag']}>{i},0,"
        f"INDEX($B${rev}:${get_column_letter(MONTHS + 1)}${rev},1,{i + 1}-{refs['collect_lag']}))"
    ))
    cout = row
    row = r(row, "เงินสดจ่าย", lambda col, i:
            f"={col}{cogs}+{col}{mkt}+{col}{oth}+{col}{fx}")
    ncf = row
    row = r(row, "กระแสเงินสดสุทธิ", lambda col, i: (
        f"={col}{cin}+{col}{cout}-{refs['capex']}" if i == 0 else f"={col}{cin}+{col}{cout}"
    ), _BAHT, True)
    cum = row
    row = r(row, "เงินสดคงเหลือสะสม", lambda col, i: (
        f"={refs['cash0']}+{col}{ncf}" if i == 0
        else f"={get_column_letter(1 + i)}{cum}+{col}{ncf}"
    ), _BAHT, True, True)

    # ── decision metrics: the numbers the CEO actually acts on ──
    row += 1
    ws.cell(row=row, column=1, value="ตัวชี้วัดสำหรับตัดสินใจ").font = Font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_CYAN)
    ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFFFF")
    row += 1
    last_col = get_column_letter(MONTHS + 1)
    metrics = [
        ("กำไรสุทธิรวม 24 เดือน", f"=SUM(B{ni}:{last_col}{ni})", _BAHT),
        ("เงินสดต่ำสุดที่เคยแตะ", f"=MIN(B{cum}:{last_col}{cum})", _BAHT),
        ("เดือนแรกที่กำไรเป็นบวก",
         f'=IFERROR(MATCH(TRUE,INDEX(B{ni}:{last_col}{ni}>0,0),0),"ไม่ถึงจุดคุ้มทุนใน 24 เดือน")', "0"),
        ("เดือนแรกที่เงินสดสะสมเป็นบวก",
         f'=IFERROR(MATCH(TRUE,INDEX(B{cum}:{last_col}{cum}>0,0),0),"เงินสดไม่กลับมาบวกใน 24 เดือน")', "0"),
        ("Runway (เดือนที่เงินสดหมด)",
         f'=IFERROR(MATCH(TRUE,INDEX(B{cum}:{last_col}{cum}<0,0),0),"เงินสดไม่ติดลบใน 24 เดือน")', "0"),
        ("อัตรากำไรสุทธิเฉลี่ย", f"=IFERROR(AVERAGE(B{margin}:{last_col}{margin}),0)", _PCT),
        ("ROI จากเงินลงทุน",
         f"=IFERROR(SUM(B{ni}:{last_col}{ni})/{refs['capex']},\"ไม่มีเงินลงทุนตั้งต้น\")", _PCT),
    ]
    first_metric = row
    for label_m, formula, fmt in metrics:
        ws.cell(row=row, column=1, value=label_m).font = Font(size=10)
        ws.cell(row=row, column=1).border = _BOX
        c = ws.cell(row=row, column=2, value=formula)
        c.number_format = fmt
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.border = _BOX
        row += 1

    # profit + cash trend chart
    chart = LineChart()
    chart.title = f"แนวโน้มกำไรและเงินสด — {label}"
    chart.height, chart.width = 8, 28
    for src_row in (ni, cum):  # series names come from column A via titles_from_data
        ref = Reference(ws, min_col=1, max_col=MONTHS + 1, min_row=src_row, max_row=src_row)
        chart.add_data(ref, from_rows=True, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=2, max_col=MONTHS + 1, min_row=5, max_row=5))
    ws.add_chart(chart, f"A{row + 1}")

    ws.column_dimensions["A"].width = 34
    for i in range(MONTHS):
        ws.column_dimensions[get_column_letter(2 + i)].width = 13
    ws.freeze_panes = "B6"
    return {"rev": rev, "ni": ni, "cum": cum, "margin": margin,
            "metrics_at": first_metric, "sheet": label}


def _compare_sheet(wb: Workbook, marks: dict, refs: dict) -> None:
    ws = wb.create_sheet("เปรียบเทียบ Scenario")
    _title(ws, "เปรียบเทียบ 3 Scenario เคียงข้างกัน", 5)
    ws["A3"] = "ตัวเลขทั้งหมดอ้างอิงสดจากชีต Base / Best / Worst"
    ws["A3"].font = Font(italic=True, size=9, color="FF57606A")

    labels = ["กำไรสุทธิรวม 24 เดือน", "เงินสดต่ำสุดที่เคยแตะ", "เดือนแรกที่กำไรเป็นบวก",
              "เดือนแรกที่เงินสดสะสมเป็นบวก", "Runway (เดือนที่เงินสดหมด)",
              "อัตรากำไรสุทธิเฉลี่ย", "ROI จากเงินลงทุน"]
    fmts = [_BAHT, _BAHT, "0", "0", "0", _PCT, _PCT]
    _hdr(ws, 5, ["ตัวชี้วัด", "Base", "Best", "Worst", "ช่วงห่าง Best-Worst"])
    for i, (lab, fmt) in enumerate(zip(labels, fmts)):
        row = 6 + i
        ws.cell(row=row, column=1, value=lab).font = Font(size=10)
        for j, (key, sheet_label, fill) in enumerate(_SCENARIOS):
            m = marks[key]
            src = f"'{sheet_label}'!$B${m['metrics_at'] + i}"
            c = ws.cell(row=row, column=2 + j, value=f"={src}")
            c.number_format = fmt
            c.fill = PatternFill("solid", fgColor=fill)
            c.font = Font(size=10)
        if fmt in (_BAHT, _PCT):
            g = ws.cell(row=row, column=5, value=f"=IFERROR(C{row}-D{row},\"—\")")
            g.number_format = fmt
            g.font = Font(bold=True, size=10)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = _BOX

    bar = BarChart()
    bar.title = "กำไรสุทธิรวม 24 เดือน ต่อ Scenario"
    bar.height, bar.width = 8, 16
    bar.add_data(Reference(ws, min_col=2, max_col=4, min_row=6, max_row=6), from_rows=True)
    bar.set_categories(Reference(ws, min_col=2, max_col=4, min_row=5, max_row=5))
    ws.add_chart(bar, "A15")

    line = LineChart()
    line.title = "เงินสดคงเหลือสะสม — ทั้ง 3 Scenario"
    line.height, line.width = 8, 22
    for key, sheet_label, _f in _SCENARIOS:
        cum = marks[key]["cum"]
        line.add_data(Reference(wb[sheet_label], min_col=1, max_col=MONTHS + 1,
                                min_row=cum, max_row=cum), from_rows=True, titles_from_data=True)
    ws.add_chart(line, "J15")

    ws.column_dimensions["A"].width = 34
    for col in "BCDE":
        ws.column_dimensions[col].width = 18


def _risk_sheet(wb: Workbook, model: dict, marks: dict, refs: dict) -> None:
    ws = wb.create_sheet("ความเสี่ยง")
    _title(ws, "การบริหารความเสี่ยง — Sensitivity, Break-even, สัญญาณเตือน", 6)

    base = marks["base"]
    ws["A3"] = "ตาราง Sensitivity: กำไรสุทธิรวม 24 เดือน เมื่อการเติบโตและต้นทุนขายเปลี่ยน"
    ws["A3"].font = Font(bold=True, size=11)
    sens = (model.get("sensitivity") or {})
    ws["A4"] = f"ตัวแปรชี้เป็นชี้ตาย: {sens.get('why', 'อัตราเติบโต vs ต้นทุนขาย')}"
    ws["A4"].font = Font(italic=True, size=9, color="FF57606A")

    # 2-way table. Excel's native Data Table needs a recalc engine we don't
    # have offline, so each cell is an explicit closed-form formula instead —
    # same result, and it survives being opened in Numbers/LibreOffice.
    growths = [-0.02, -0.01, 0.0, 0.01, 0.02, 0.03, 0.05]
    cogs_deltas = [-0.10, -0.05, 0.0, 0.05, 0.10]
    top = 6
    ws.cell(row=top, column=1, value="Δ ต้นทุนขาย \\ Δ การเติบโต").font = Font(bold=True, size=9)
    ws.cell(row=top, column=1).fill = PatternFill("solid", fgColor=_NAVY)
    ws.cell(row=top, column=1).font = Font(bold=True, size=9, color="FFFFFFFF")
    ws.cell(row=top, column=1).border = _BOX
    for j, g in enumerate(growths):
        c = ws.cell(row=top, column=2 + j, value=g)
        c.number_format = '+0.0%;-0.0%;0.0%'
        c.font = Font(bold=True, size=9, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=_CYAN)
        c.border = _BOX
    for i, d in enumerate(cogs_deltas):
        row = top + 1 + i
        rc = ws.cell(row=row, column=1, value=d)
        rc.number_format = '+0.0%;-0.0%;0.0%'
        rc.font = Font(bold=True, size=9, color="FFFFFFFF")
        rc.fill = PatternFill("solid", fgColor=_CYAN)
        rc.border = _BOX
        for j, g in enumerate(growths):
            gl = get_column_letter(2 + j)
            # units_m1*price*Σ(1+growth+Δg)^t * (1 - cogs-Δc - mkt - other) - fixed*24
            f = (f"={refs['units_m1']}*{refs['price']}*{refs['base_rev_mult']}"
                 f"*SUMPRODUCT((1+{refs['growth']}+{gl}${top}-{refs['churn']})"
                 f"^(ROW(INDIRECT(\"1:{MONTHS}\"))-1))"
                 f"*(1-({refs['cogs_pct']}+$A{row})*{refs['base_cost_mult']}"
                 f"-{refs['mkt_pct']}*{refs['base_cost_mult']}"
                 f"-{refs['other_pct']}*{refs['base_cost_mult']})"
                 f"-{refs['fixed_cost']}*{refs['base_cost_mult']}*{MONTHS}")
            c = ws.cell(row=row, column=2 + j, value=f)
            c.number_format = _BAHT
            c.font = Font(size=9)
            c.border = _BOX

    # break-even: how many units/month to cover fixed cost at current margins
    row = top + len(cogs_deltas) + 3
    ws.cell(row=row, column=1, value="จุดคุ้มทุน (Break-even)").font = Font(bold=True, size=12)
    row += 1
    cm = (f"({refs['price']}*(1-({refs['cogs_pct']}+{refs['mkt_pct']}+{refs['other_pct']})"
          f"*{refs['base_cost_mult']}))")
    be = [
        ("กำไรส่วนเกินต่อหน่วย (Contribution/unit)", f"={cm}", _BAHT),
        ("หน่วยที่ต้องขายต่อเดือนถึงจะคุ้มทุน",
         f"=IFERROR({refs['fixed_cost']}*{refs['base_cost_mult']}/{cm},\"ขาดทุนทุกหน่วย\")", "#,##0.0"),
        ("รายได้ต่อเดือนที่ต้องทำถึงจะคุ้มทุน",
         f"=IFERROR({refs['fixed_cost']}*{refs['base_cost_mult']}/{cm}*{refs['price']},\"—\")", _BAHT),
        ("ส่วนต่างความปลอดภัย (Margin of Safety) เดือน 1",
         f"=IFERROR(('Base'!B{base['rev']}-{refs['fixed_cost']}*{refs['base_cost_mult']}"
         f"/{cm}*{refs['price']})/'Base'!B{base['rev']},\"—\")", _PCT),
        ("เดือนที่ต้องเฝ้าที่สุด (เงินสดต่ำสุด)",
         f"=IFERROR(MATCH(MIN('Base'!B{base['cum']}:{get_column_letter(MONTHS + 1)}{base['cum']}),"
         f"'Base'!B{base['cum']}:{get_column_letter(MONTHS + 1)}{base['cum']},0),\"—\")", "0"),
    ]
    for lab, f, fmt in be:
        ws.cell(row=row, column=1, value=lab).font = Font(size=10)
        ws.cell(row=row, column=1).border = _BOX
        c = ws.cell(row=row, column=2, value=f)
        c.number_format = fmt
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=_WARN)
        c.border = _BOX
        row += 1

    # risk register straight from the CFO's assumptions
    row += 2
    ws.cell(row=row, column=1, value="ทะเบียนความเสี่ยง + สัญญาณเตือน").font = Font(bold=True, size=12)
    row += 1
    _hdr(ws, row, ["ความเสี่ยง", "ตัวแปรในแบบจำลอง", "สัญญาณเตือน (เชิงตัวเลข)", "แผนรับมือ"])
    row += 1
    risks = model.get("risks") or []
    for rk in risks[:12]:
        if not isinstance(rk, dict):
            continue
        for col, key in enumerate(("risk", "driver", "trigger", "mitigation"), start=1):
            c = ws.cell(row=row, column=col, value=str(rk.get(key) or "—"))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=9)
            c.border = _BOX
        ws.row_dimensions[row].height = 34
        row += 1
    if not risks:
        ws.cell(row=row, column=1, value="(CFO ไม่ได้ระบุความเสี่ยงเชิงตัวเลข)").font = Font(
            italic=True, size=9, color="FF57606A")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="KPI ที่ต้องดูทุกเดือน").font = Font(bold=True, size=12)
    row += 1
    for k in (model.get("kpis_to_watch") or ["—"])[:12]:
        ws.cell(row=row, column=1, value=f"• {k}").font = Font(size=10)
        row += 1

    ws.column_dimensions["A"].width = 40
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 20


def _howto_sheet(wb: Workbook, session: dict, model: dict, provider: str) -> None:
    ws = wb.create_sheet("วิธีอ่าน", 0)
    _title(ws, "วิธีใช้แบบจำลองนี้ + ที่มาของตัวเลข", 3)
    lines = [
        ("คำถามของ CEO", session.get("question", "—")),
        ("ธุรกิจนี้หาเงินจาก", model.get("business_model", "—")),
        ("หน่วยเงิน / ช่วงเวลา", model.get("currency_note", "บาท (THB), 24 เดือน")),
        ("CFO ที่สร้างแบบจำลอง", f"{provider} · consult #{session.get('id')}"),
        ("", ""),
        ("วิธีใช้ (สำคัญ)", ""),
        ("1", "แก้ตัวเลขได้เฉพาะช่องสีเหลืองในชีต 'สมมติฐาน' — ทุกชีตอื่นเป็นสูตรที่คำนวณต่อเอง"),
        ("2", "เปิดชีต 'เปรียบเทียบ Scenario' เพื่อดูผลกระทบทั้ง 3 กรณีพร้อมกัน"),
        ("3", "ชีต 'ความเสี่ยง' บอกจุดคุ้มทุน เดือนที่เงินสดตึงที่สุด และสัญญาณเตือนที่ต้องเฝ้า"),
        ("4", "คอลัมน์ 'ประเภท' ในชีตสมมติฐานบอกว่าเลขไหนมาจากบทถกเถียงจริง เลขไหน CFO ประมาณเอง"),
        ("", ""),
        ("ข้อจำกัดที่ต้องรู้", ""),
        ("•", "ช่องที่ระบุ 'ประมาณการ' คือตัวเลขที่ไม่มีข้อมูลจริงรองรับ — ต้องแทนที่ด้วยตัวเลขจริงของคุณก่อนใช้ตัดสินใจ"),
        ("•", "แบบจำลองนี้ไม่รวมภาษี ค่าเสื่อมราคา ดอกเบี้ย และเงินทุนหมุนเวียนอื่นนอกจากระยะเวลาเก็บเงิน"),
        ("•", "การเติบโตใช้อัตราทบต้นคงที่ ธุรกิจจริงมักไม่โตเป็นเส้นเรียบ — ใช้ Worst case ถ่วงน้ำหนักเสมอ"),
        ("•", "ตัวเลขนี้คือกรอบคิดเพื่อตั้งคำถาม ไม่ใช่คำพยากรณ์"),
    ]
    row = 3
    for k, v in lines:
        a = ws.cell(row=row, column=1, value=k)
        a.font = Font(bold=not v or k in ("วิธีใช้ (สำคัญ)", "ข้อจำกัดที่ต้องรู้"), size=10)
        b = ws.cell(row=row, column=2, value=v)
        b.font = Font(size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30 if len(str(v)) > 70 else 18
        row += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96


# ── public API ──

def assumptions(session: dict, refresh: bool = False) -> dict:
    """Ask the CFO's provider for model inputs. Cached on the session."""
    cached = session.get("finmodel")
    if not refresh and isinstance(cached, dict) and cached:
        return cached

    transcript = []
    for step in session.get("steps") or []:
        for key, val in (step.get("results") or {}).items():
            who = config.DEPTS.get(key, {}).get("name", key)
            transcript.append(f"[{who}]\n{val.get('text', '')}")
    if not transcript:
        return {}

    provider = store.get_providers().get("cfo", "mock")
    out = llm.chat(provider, ASSUMPTION_SYSTEM,
                   f"คำถามของ CEO: {session['question']}\n\nบทถกเถียงของบอร์ด:\n"
                   + "\n\n".join(transcript), max_tokens=8192)
    if not out["ok"]:
        log.warning("finmodel assumptions failed for consult %s: provider not ok",
                    session.get("id"))
        return {}
    from .report import _parse_json  # local import avoids a cycle at module load
    data = _parse_json(out["text"])
    if not isinstance(data, dict) or not data:
        log.warning("finmodel assumptions unparseable for consult %s (chars=%s)",
                    session.get("id"), len(out.get("text") or ""))
        return {}
    data["_provider"] = out.get("provider", provider)
    store.attach_finmodel(session["id"], data)
    return data


def build_workbook(session: dict) -> bytes:
    """Render the scenario model. Raises ValueError when there's nothing to model."""
    model = assumptions(session)
    if not model:
        raise ValueError("ยังไม่มีสมมติฐานการเงิน — รันบอร์ดให้ถึงรอบความเห็น (Round 1) ก่อน")

    wb = Workbook()
    wb.remove(wb.active)
    refs = _assumptions_sheet(wb, model)
    marks = {key: _scenario_sheet(wb, key, label, refs, fill)
             for key, label, fill in _SCENARIOS}
    _compare_sheet(wb, marks, refs)
    _risk_sheet(wb, model, marks, refs)
    _howto_sheet(wb, session, model, str(model.get("_provider", "—")))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
