"""Hub tests — stepwise board flow, decision gates, branch/reset, web research, PDF reports,
CFO Excel scenario model."""
import io
import json
import os
import tempfile
from pathlib import Path
from unittest.mock import patch

import httpx
import pytest
from fastapi.testclient import TestClient

from app import config, deliverable


@pytest.fixture()
def client(tmp_path, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "MEMORY_DIR", tmp_path)
    import app.store as store
    monkeypatch.setattr(store, "_FILE", tmp_path / "hub_store.json")
    store._STOP.clear()
    import app.docs as docs
    monkeypatch.setattr(docs, "LOCAL_DIR", tmp_path)
    monkeypatch.setattr(docs, "_META", tmp_path / "_meta.json")
    import app.depts as depts
    depts._health_cache.clear()
    monkeypatch.setattr(depts, "svc_health", lambda dept: False)
    from app.main import app
    return TestClient(app)


def _fake_chat(provider, system, user, cancel=None, **kw):
    # **kw absorbs per-call knobs like max_tokens so a fake never has to track
    # llm.chat's signature to stay usable.
    return {"text": f"[{provider}] คำตอบทดสอบ", "provider": provider, "model": "m", "ok": True}


FAKE_SOURCES = [
    {"title": "ตลาดตู้กดอัตโนมัติไทย 2026", "url": "https://example.com/market",
     "snippet": "มูลค่าตลาด 4,200 ล้านบาท", "body": "มูลค่าตลาด 4,200 ล้านบาท เติบโต 12%",
     "query": "ขนาดตลาดตู้กดอัตโนมัติ"},
    {"title": "Vending machine competition", "url": "https://example.com/comp",
     "snippet": "average ticket 45 THB", "body": "average ticket 45 THB",
     "query": "vending competitors thailand"},
]


def _start(client, question="ควรเปิดสาขาใหม่ไหม", project=None, web=False):
    r = client.post("/api/consult", json={"question": question, "project": project,
                                          "web_research": web})
    assert r.status_code == 200
    return r.json()


def _advance(client, sid, step=None, directive=None):
    r = client.post(f"/api/consult/{sid}/advance", json={"step": step, "directive": directive})
    assert r.status_code == 200, r.text
    return r.json()


def _frame_json(seats=None):
    return json.dumps({
        "reframed": "ควรผูกสัญญาเช่าสาขาใหม่ 3 ปีในไตรมาสนี้หรือไม่",
        "decision_type": "กลับตัวยาก — ผูกสัญญาเช่า 3 ปี",
        "seats": list(seats if seats is not None else config.DEPTS),
        "seat_reasons": {"cfo": "เป็นการผูกพันเงินสดระยะยาว"},
        "excluded": {},
        "what_would_change_the_answer": ["ยอดขายต่อสาขาจริงย้อนหลัง 6 เดือน"],
        "success_criteria": ["สาขาใหม่คุ้มทุนภายใน 14 เดือน"],
        "framing_risk": "",
    }, ensure_ascii=False)


def _reply(text, provider="anthropic"):
    return {"text": text, "provider": provider, "model": "m", "ok": True}


def _framed(client, seats=None, **kw):
    """Open a session and run Stage 1, so the board is convened and the later
    stages have a framing to argue inside."""
    s = _start(client, **kw)
    with patch("app.llm.chat", return_value=_reply(_frame_json(seats))):
        return _advance(client, s["id"])


# ── state ──

def test_state_shape(client):
    s = client.get("/api/state").json()
    assert {d["key"] for d in s["depts"]} == set(config.DEPTS)
    assert {p["key"] for p in s["providers"]} >= {"anthropic", "gemini", "manus", "mock"}
    assert set(s["decision_stats"]) == {"total", "scored", "saved", "faster", "neutral", "missed"}
    # department service health is surfaced so the UI can show the online dot
    assert all("online" in d for d in s["depts"])
    # web research advertises its state; the suite runs with the keys cleared
    assert s["research"]["backend"] == "none" and s["research"]["label"]
    assert s["research"]["keyed"] is False


# ── stepwise board flow with decision gates ──

def test_consult_stops_at_a_gate_before_every_stage(client):
    s = _start(client)
    # creating a session runs nothing — the CEO owns the first move, and Frame
    # is the first thing that happens, before anyone is even convened
    assert s["steps"] == [] and s["next_step"] == "frame" and s["status"] == "awaiting"

    with patch("app.llm.chat", return_value=_reply(_frame_json())) as m:
        s = _advance(client, s["id"])
    assert [x["key"] for x in s["steps"]] == ["frame"]
    assert s["next_step"] == "positions" and s["status"] == "awaiting"
    assert m.call_count == 1                      # one moderator, not a whole board

    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        s = _advance(client, s["id"])
    assert [x["key"] for x in s["steps"]] == ["frame", "positions"]
    assert s["next_step"] == "debate"
    assert m.call_count == len(config.DEPTS)      # one call per convened seat

    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])             # debate
        s = _advance(client, s["id"])             # red-team & converge
    assert [x["key"] for x in s["steps"]] == ["frame", "positions", "debate", "redteam"]
    assert s["next_step"] == "brief"
    # the red team is an outside voice, on top of every seat's own self-audit
    assert "redteam" in s["steps"][-1]["results"]

    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
    assert s["next_step"] is None and s["status"] == "done"
    assert s["steps"][-1]["results"]["chair"]["text"]


def test_ceo_directive_is_injected_into_the_round(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        s = _advance(client, s["id"], directive="เน้นความเสี่ยงกระแสเงินสดเป็นหลัก")
    assert all("เน้นความเสี่ยงกระแสเงินสดเป็นหลัก" in c.args[2] for c in m.call_args_list)
    assert s["steps"][-1]["directive"] == "เน้นความเสี่ยงกระแสเงินสดเป็นหลัก"


def test_ceo_can_skip_straight_to_the_brief(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"], step="brief")
    assert [x["key"] for x in s["steps"]] == ["frame", "brief"]
    assert s["next_step"] is None and s["status"] == "done"   # the brief is terminal


def test_rerunning_a_finished_stage_is_rejected(client):
    s = _framed(client)
    r = client.post(f"/api/consult/{s['id']}/advance", json={"step": "frame"})
    assert r.status_code == 400 and "reset" in r.text


# ── branch & reset (git-style history the board learns from) ──

def test_reset_rewinds_and_feeds_the_rejected_path_back(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])            # positions
        s = _advance(client, s["id"])            # debate
    assert len(s["steps"]) == 3

    r = client.post(f"/api/consult/{s['id']}/reset", json={"step": "debate"})
    s = r.json()
    assert [x["key"] for x in s["steps"]] == ["frame", "positions"]
    assert [h["key"] for h in s["history"]] == ["debate"]
    assert s["history"][0]["reason"] == "reset" and s["next_step"] == "debate"

    # the discarded round is replayed so the board must find a different angle
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    assert all("ห้ามเสนอซ้ำแนวเดิม" in c.args[2] for c in m.call_args_list)


def test_branch_forks_an_alternate_timeline_keeping_the_original(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
        s = _advance(client, s["id"])

    child = client.post(f"/api/consult/{s['id']}/branch", json={"step": "debate"}).json()
    assert child["id"] != s["id"] and child["parent_id"] == s["id"]
    assert child["branched_from"] == "debate"
    assert [x["key"] for x in child["steps"]] == ["frame", "positions"]  # shared prefix
    assert child["history"][0]["reason"] == "branch"          # remembers the road not taken
    assert child["seats"] == s["seats"]                        # same board, different route

    original = client.get(f"/api/consults/{s['id']}").json()
    assert [x["key"] for x in original["steps"]] == ["frame", "positions", "debate"]


def test_branch_on_a_stage_that_never_ran_is_rejected(client):
    s = _start(client)
    r = client.post(f"/api/consult/{s['id']}/branch", json={"step": "redteam"})
    assert r.status_code == 400
    assert client.post(f"/api/consult/{s['id']}/branch", json={"step": "nope"}).status_code == 400


# ── stop ──

def test_stop_marks_the_session_and_abandons_pending_advisors(client):
    s = _framed(client)
    sid = s["id"]

    def chat_then_stop(provider, system, user, cancel=None, **kw):
        client.post(f"/api/consult/{sid}/stop")   # CEO hits STOP mid-round
        return _fake_chat(provider, system, user)

    with patch("app.llm.chat", side_effect=chat_then_stop) as m:
        s = _advance(client, sid)
    assert s["status"] == "stopped"
    # every seat holds a live cancel probe, so a slow provider (Manus polls for
    # minutes) bails out instead of waiting out its own deadline
    probes = [c.kwargs["cancel"] for c in m.call_args_list]
    assert probes and all(callable(p) for p in probes)
    # partial results are kept so nothing the board already said is lost
    assert s["steps"][-1]["key"] == "positions"
    assert len(s["steps"][-1]["results"]) == len(config.DEPTS)

    # and the CEO can resume from where it stopped
    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, sid)
    assert s["status"] == "awaiting" and len(s["steps"]) == 3


def test_stop_on_unknown_session_404s(client):
    assert client.post("/api/consult/999/stop").status_code == 404


# ── Stage 1: frame ──

def test_frame_convenes_only_the_seats_the_question_needs(client):
    s = _start(client)
    with patch("app.llm.chat", return_value=_reply(_frame_json(["cfo", "coo"]))):
        s = _advance(client, s["id"])
    framer = s["steps"][0]["results"]["framer"]
    assert framer["ok"] and framer["seats"] == ["cfo", "coo"]
    assert s["seats"] == ["cfo", "coo"] and s["convened"] == ["cfo", "coo"]
    assert framer["reframed"].startswith("ควรผูกสัญญาเช่า")

    # the uninvited seats never speak — calling everyone is noise, not rigour
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        s = _advance(client, s["id"])
    assert set(s["steps"][-1]["results"]) == {"cfo", "coo"}
    assert m.call_count == 2


def test_a_one_voice_board_is_refused(client):
    """A debate needs someone to disagree; one seat is a monologue."""
    s = _start(client)
    with patch("app.llm.chat", return_value=_reply(_frame_json(["cfo"]))):
        s = _advance(client, s["id"])
    assert s["seats"] == list(config.DEPTS)


def test_frame_failure_still_convenes_the_board(client):
    s = _start(client)
    with patch("app.llm.chat", return_value=_reply("ขอโทษครับ ผมไม่เข้าใจ")):
        s = _advance(client, s["id"])
    framer = s["steps"][0]["results"]["framer"]
    assert framer["ok"] is False and framer["seats"] == list(config.DEPTS)
    assert s["next_step"] == "positions"          # the session is not blocked


def test_the_framing_travels_into_every_later_stage(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    for c in m.call_args_list:
        assert "กรอบการตัดสินใจที่ moderator ตั้งไว้" in c.args[2]
        assert "ควรผูกสัญญาเช่าสาขาใหม่ 3 ปีในไตรมาสนี้หรือไม่" in c.args[2]


# ── Stage 2: independent research, one desk per seat ──

def test_every_seat_researches_independently(client):
    s = _framed(client, question="ควรลงทุนตู้กดอัตโนมัติไหม", web=True)
    assert s["next_step"] == "research"

    with patch("app.research.gather", return_value={"sources": FAKE_SOURCES, "errors": []}) as g, \
         patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
    research_step = s["steps"][-1]
    assert research_step["key"] == "research"

    # one desk per seat — shared research would hand every seat the same
    # evidence, and with it the same blind spot
    for dept in config.DEPTS:
        desk = research_step["results"][dept]
        assert desk["ok"] and desk["queries"]
        assert [x["url"] for x in desk["sources"]] == [x["url"] for x in FAKE_SOURCES]
    assert g.call_count == len(config.DEPTS)

    # the merged index is the union, tagged with which seat found what
    merged = research_step["results"]["analyst"]
    assert {x["url"] for x in merged["sources"]} == {x["url"] for x in FAKE_SOURCES}
    assert all(x["found_by"] in config.DEPTS for x in merged["sources"])

    # downstream, a seat argues from the evidence IT found
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    for c in m.call_args_list:
        assert "หลักฐานที่คุณค้นมาเอง" in c.args[2]
        assert "https://example.com/market" in c.args[2]


def test_each_seat_writes_its_own_queries_on_its_own_model(client):
    s = _framed(client, web=True)
    reply = _reply("ขนาดตลาดตู้กดไทย\nvending machine market thailand\nกฎหมายตู้หยอดเหรียญ")
    with patch("app.llm.chat", return_value=reply) as m, \
         patch("app.research.gather", return_value={"sources": FAKE_SOURCES, "errors": []}) as g:
        _advance(client, s["id"])
    assert g.call_args.args[0] == ["ขนาดตลาดตู้กดไทย", "vending machine market thailand",
                                   "กฎหมายตู้หยอดเหรียญ"]
    # the query prompt is the seat's own, in its own lane, on its assigned agent
    query_calls = [c for c in m.call_args_list if "ตั้งคำค้นสำหรับ search engine" in c.args[1]]
    assert len(query_calls) == len(config.DEPTS)
    assert {c.args[0] for c in query_calls} == {
        config.DEFAULT_PROVIDERS.get(d, "mock") for d in config.DEPTS}


def test_research_failure_degrades_to_documents_only(client):
    s = _framed(client, web=True)
    with patch("app.research.gather", return_value={"sources": [], "errors": []}), \
         patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
    desks = s["steps"][-1]["results"]
    assert all(desks[d]["ok"] is False and desks[d]["sources"] == [] for d in config.DEPTS)
    assert desks["analyst"]["ok"] is False

    # the board still debates — a dead search backend cannot block the consult
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        s = _advance(client, s["id"])
    assert s["steps"][-1]["key"] == "positions"
    assert all("หลักฐานที่คุณค้นมาเอง" not in c.args[2] for c in m.call_args_list)


def test_web_research_can_be_turned_off(client):
    s = _start(client, web=False)
    assert s["next_step"] == "frame"          # framing always happens
    with patch("app.llm.chat", return_value=_reply(_frame_json())):
        s = _advance(client, s["id"])
    assert s["next_step"] == "positions"      # the research stage is skipped


def test_no_key_means_no_search_not_a_quiet_internet(client, monkeypatch):
    """With no key the board must refuse to search and say so. Returning [] here
    is what let an unsearchable board hand the CEO an evidence-free brief."""
    from app import config, research
    for k in ("TAVILY_API_KEY", "SERPAPI_API_KEY", "BRAVE_API_KEY", "SERPER_API_KEY"):
        monkeypatch.setattr(config, k, "")
    assert research.backend() == "none" and research.configured() == []
    out = research.search_detail("ตลาดบาร์ เอกมัย")
    assert out["results"] == [] and out["engine"] is None
    assert "API key" in out["error"], out["error"]
    assert "TAVILY_API_KEY" in out["error"], "the error must name the fix"


def test_backend_priority_and_label_follow_the_keys(client, monkeypatch):
    from app import config, research
    for k in ("TAVILY_API_KEY", "SERPAPI_API_KEY", "BRAVE_API_KEY", "SERPER_API_KEY"):
        monkeypatch.setattr(config, k, "")
    for env, name in (("SERPER_API_KEY", "serper"), ("BRAVE_API_KEY", "brave"),
                      ("SERPAPI_API_KEY", "serpapi"), ("TAVILY_API_KEY", "tavily")):
        monkeypatch.setattr(config, env, "x")
        assert research.backend() == name, f"{env} should take the top slot"
    assert "Tavily" in research.backend_label() and "+3" in research.backend_label()


def test_search_fails_over_across_keyed_backends(client, monkeypatch):
    """Tavily rate-limiting for a minute must cost latency, not the evidence."""
    from app import config, research
    monkeypatch.setattr(config, "TAVILY_API_KEY", "x")
    monkeypatch.setattr(config, "SERPAPI_API_KEY", "y")
    monkeypatch.setattr(config, "BRAVE_API_KEY", "")
    monkeypatch.setattr(config, "SERPER_API_KEY", "")
    hits = [{"title": "t", "url": "https://example.com/a", "snippet": "s"}]
    rate_limited = httpx.HTTPStatusError(
        "429", request=httpx.Request("POST", "https://api.tavily.com"),
        response=httpx.Response(429))
    with patch("app.research._tavily", side_effect=rate_limited), \
         patch("app.research._serpapi", return_value=hits) as second:
        out = research.search_detail("ตลาดบาร์ เอกมัย")
    assert second.called, "the next keyed backend must get a turn"
    assert out["results"] == hits and out["engine"] == "serpapi" and out["error"] is None


def test_every_keyed_backend_failing_reports_each_reason(client, monkeypatch):
    """The bug this guards: a 403 and a genuinely empty index both arrived as an
    empty list, so the board reported no evidence for a search that never ran."""
    from app import config, research
    monkeypatch.setattr(config, "TAVILY_API_KEY", "x")
    monkeypatch.setattr(config, "SERPAPI_API_KEY", "y")
    monkeypatch.setattr(config, "BRAVE_API_KEY", "")
    monkeypatch.setattr(config, "SERPER_API_KEY", "")
    blocked = httpx.HTTPStatusError(
        "403", request=httpx.Request("GET", "https://api.tavily.com"),
        response=httpx.Response(403))
    with patch("app.research._tavily", side_effect=blocked), \
         patch("app.research._serpapi", side_effect=RuntimeError("down")):
        out = research.search_detail("ตลาดบาร์ เอกมัย")
    assert out["results"] == [] and out["engine"] is None
    assert "tavily" in out["error"] and "serpapi" in out["error"]
    assert "403" in out["error"]


def test_search_never_raises_when_every_backend_dies(client, monkeypatch):
    from app import config, research
    monkeypatch.setattr(config, "TAVILY_API_KEY", "x")
    for k in ("SERPAPI_API_KEY", "BRAVE_API_KEY", "SERPER_API_KEY"):
        monkeypatch.setattr(config, k, "")
    with patch("app.research._tavily", side_effect=RuntimeError("boom")):
        assert research.search("อะไรก็ได้") == []


def test_the_scraped_engines_are_gone(client):
    """Keyless scrapers answered bot checks that parsed as zero results. They
    must not come back through a helper left behind in the module."""
    from app import research
    for gone in ("_duckduckgo", "_mojeek", "_parse_ddg", "_unwrap_ddg",
                 "_DDG_ENDPOINTS", "_DDG_HEADERS", "_KEYLESS_CHAIN", "_BACKENDS"):
        assert not hasattr(research, gone), f"{gone} is still present"
    src = (Path(__file__).resolve().parent.parent / "app" / "research.py").read_text("utf-8")
    body = src.split('"""', 2)[-1]      # past the docstring that explains why they went
    assert "duckduckgo" not in body.lower() and "mojeek" not in body.lower()
    assert [n for n, _e, _f, _l in research._PRIORITY] == \
        ["tavily", "serpapi", "brave", "serper"]


def test_every_priority_entry_has_a_caller_a_key_and_a_label(client):
    """A backend that resolves but has no implementation would silently return
    nothing, which reads as the internet having no data on the subject."""
    from app import research
    for name, env, fn, label in research._PRIORITY:
        assert hasattr(config, env), f"config is missing {env}"
        assert hasattr(research, fn), f"{name} has no caller"
        assert label, f"{name} has no label"


def test_serpapi_reads_organic_results_and_defaults_to_thai(client, monkeypatch):
    """The engine=google endpoint returns organic_results with title/link/snippet.
    Thai locale is the sensible default here; the CEO's business is in Thailand,
    and a Thai query landing on the US SERP is the failure mode to prevent."""
    from app import config, research
    monkeypatch.setattr(config, "SERPAPI_API_KEY", "serpapi-test")
    monkeypatch.setattr(config, "TAVILY_API_KEY", "")   # SerpApi is now first
    assert research.backend() == "serpapi"
    payload = {"organic_results": [
        {"title": "บาร์เอกมัย 2026", "link": "https://example.com/bars",
         "snippet": "รวมบาร์เปิดใหม่ ราคาเฉลี่ย 450 บาท"}]}

    class _Resp:
        def json(self):
            return payload

        def raise_for_status(self):
            pass

    with patch("httpx.get", return_value=_Resp()) as get:
        out = research._serpapi("บาร์ เอกมัย", 5)
    params = get.call_args.kwargs["params"]
    assert params["engine"] == "google" and params["hl"] == "th" and params["gl"] == "th"
    assert params["api_key"] == "serpapi-test"
    assert out[0]["url"] == "https://example.com/bars"
    assert "450 บาท" in out[0]["snippet"]


def test_serpapi_surfaces_plan_or_quota_errors_carried_in_200_ok(client, monkeypatch):
    """SerpApi returns HTTP 200 with an "error" field when the plan is out — a
    caller that only checks the status code silently treats it as empty results."""
    from app import config, research
    monkeypatch.setattr(config, "SERPAPI_API_KEY", "x")

    class _Resp:
        def json(self):
            return {"error": "You have reached your monthly search limit"}

        def raise_for_status(self):
            pass

    with patch("httpx.get", return_value=_Resp()):
        with pytest.raises(RuntimeError, match="monthly search limit"):
            research._serpapi("q", 5)


def test_tavily_wins_over_serpapi_when_both_are_configured(client, monkeypatch):
    """Tavily returns extracted page bodies with the results, so the analyst can
    read the full content instead of re-fetching each URL. A key that ships more
    per request should outrank a plain SERP scraper."""
    from app import config, research
    monkeypatch.setattr(config, "TAVILY_API_KEY", "t")
    monkeypatch.setattr(config, "SERPAPI_API_KEY", "s")
    assert research.backend() == "tavily"


def test_tavily_authenticates_by_header_and_asks_for_page_bodies(client, monkeypatch):
    """Tavily's documented auth is a Bearer header, and raw_content is opt-in —
    without it every source came back empty and had to be re-fetched by hand."""
    from app import config, research
    monkeypatch.setattr(config, "TAVILY_API_KEY", "tvly-test")
    payload = {"results": [{"title": "t", "url": "https://example.com/a",
                            "content": "snippet", "raw_content": "หน้าเต็ม"}]}

    class _Resp:
        def json(self):
            return payload

        def raise_for_status(self):
            pass

    with patch("httpx.post", return_value=_Resp()) as post:
        out = research._tavily("q", 3)
    kwargs = post.call_args.kwargs
    assert kwargs["headers"]["Authorization"] == "Bearer tvly-test"
    assert kwargs["json"]["include_raw_content"] is True
    assert "api_key" not in kwargs["json"]          # the key never rides in the body
    assert out[0]["content"] == "หน้าเต็ม"           # body arrives without a second fetch


def test_gather_hands_the_reason_back_to_the_caller(client):
    from app import research
    with patch("app.research.search_detail",
               return_value={"results": [], "error": "ถูกบล็อก (HTTP 403)", "engine": None}):
        out = research.gather(["คำค้น ก", "คำค้น ข"])
    assert out["sources"] == [] and len(out["errors"]) == 2
    assert all("ถูกบล็อก" in e for e in out["errors"])


def test_a_seat_that_was_blocked_says_so_instead_of_no_evidence(client):
    s = _framed(client, web=True)
    with patch("app.research.gather",
               return_value={"sources": [], "errors": ["«q» → ถูกบล็อก (HTTP 403)"]}), \
         patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
    desk = s["steps"][-1]["results"]["cfo"]
    assert desk["blocked"] is True
    assert "ยังไม่ได้อ่านอินเทอร์เน็ตเลย" in desk["text"]
    assert "TAVILY_API_KEY" in desk["text"]          # tells the CEO the actual fix
    # the merged index warns that the board's conclusion rests on thin evidence
    index = s["steps"][-1]["results"]["analyst"]
    assert index["blocked_seats"] and "ค้นเว็บไม่ได้เลย" in index["text"]
    # and the UI must still render that index with zero sources — hiding it when
    # nothing was found buries the warning in the one case it exists for
    html = client.get("/").text
    assert "const blockedSeats = (merged.blocked_seats || []).length;" in html
    assert "const index = (found || blockedSeats)" in html


def test_a_search_that_ran_but_found_nothing_is_not_blamed_on_blocking(client):
    s = _framed(client, web=True)
    with patch("app.research.gather", return_value={"sources": [], "errors": []}), \
         patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])
    desk = s["steps"][-1]["results"]["cfo"]
    assert desk["blocked"] is False
    assert "ค้นได้แต่ไม่พบหลักฐาน" in desk["text"]
    assert "TAVILY_API_KEY" not in desk["text"]


def test_diagnose_reports_a_working_search(client, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "TAVILY_API_KEY", "x")
    hits = [{"title": "ตลาดร้านอาหาร", "url": "https://example.com/x", "snippet": "s"}]
    with patch("app.research._tavily", return_value=hits):
        j = client.get("/api/research/diagnose").json()
    assert j["ok"] is True and j["found"] == 1 and j["engine"] == "tavily"
    assert j["error"] is None and j["sample"][0]["url"] == "https://example.com/x"
    assert j["keyed"] is True and j["backends"][0] == "tavily"


def test_diagnose_reports_a_blocked_search(client, monkeypatch):
    from app import config
    monkeypatch.setattr(config, "TAVILY_API_KEY", "x")
    for k in ("SERPAPI_API_KEY", "BRAVE_API_KEY", "SERPER_API_KEY"):
        monkeypatch.setattr(config, k, "")
    with patch("app.research._tavily", side_effect=research_blocked()):
        j = client.get("/api/research/diagnose").json()
    assert j["ok"] is False and j["found"] == 0
    assert "ถูกบล็อก" in j["error"] and j["keyed"] is True


def test_diagnose_says_so_when_no_key_is_configured(client, monkeypatch):
    """An unsearchable board must be loud about it — this is the state that used
    to masquerade as "the web has nothing on this"."""
    from app import config
    for k in ("TAVILY_API_KEY", "SERPAPI_API_KEY", "BRAVE_API_KEY", "SERPER_API_KEY"):
        monkeypatch.setattr(config, k, "")
    j = client.get("/api/research/diagnose").json()
    assert j["ok"] is False and j["keyed"] is False and j["backends"] == []
    assert "API key" in j["error"]


def research_blocked():
    return httpx.HTTPStatusError("403", request=httpx.Request("GET", "https://x"),
                                 response=httpx.Response(403))


# ── PDF reports ──

METHOD_JSON = json.dumps({
    "principles": ["Unit economics ต่อจุด", "Cash runway analysis"],
    "advisors": [{"dept": "cfo", "frameworks": ["Payback period"],
                  "logic": "เทียบ payback กับเงินสดคงเหลือ",
                  "assumptions": ["ต้นทุนต่อตู้คงที่"], "limitations": ["ไม่รวมค่าซ่อม"],
                  "citations": ["เว็บ:[1]"]}],
    "data_table": {"columns": ["ตัวชี้วัด", "ค่า", "หน่วย", "ที่มา"],
                   "rows": [["payback", "14", "เดือน", "cfo"],
                            ["runway", "8", "เดือน", "cfo"]]},
    "chart": {"title": "ระยะเวลา", "labels": ["payback", "runway"],
              "values": [14, 8], "unit": "เดือน"},
}, ensure_ascii=False)

OPTIONS_JSON = json.dumps({
    "advisor_takeaways": [{"dept": "cfo", "headline": "runway ไม่พอ", "key_risk": "เงินสดขาดมือ"}],
    "options": [
        {"name": "นำร่อง 4 จุด", "summary": "ลงทุน 4 จุดก่อน", "pros": ["เสี่ยงจำกัด"],
         "cons": ["เสียทำเลบางจุด"], "supporters": ["cfo", "coo"], "opponents": [],
         "choose_when": "ถ้าต้องรักษา runway", "first_move": "ล็อกสัญญา 4 จุด"},
        {"name": "ยังไม่ขยาย", "summary": "โฟกัสจุดเดิม", "pros": ["ไม่กระทบเงินสด"],
         "cons": ["เสียโอกาส"], "supporters": ["cfo"], "opponents": ["datalyst"],
         "choose_when": "ถ้ายอดต่อจุดยังต่ำ", "first_move": "ทำโปรโมชัน"},
    ],
    "recommended": "นำร่อง 4 จุด",
    "recommended_why": "ทางเดียวที่ CFO และ COO รับได้",
    "ceo_must_decide": "จะยอมเสียทำเลเพื่อรักษา runway หรือไม่",
}, ensure_ascii=False)


def _finished_consult(client, web=False):
    s = _framed(client, question="ควรขยายสาขาไหม", web=web)
    with patch("app.llm.chat", side_effect=_fake_chat), \
         patch("app.research.gather", return_value={"sources": FAKE_SOURCES, "errors": []}):
        while s["next_step"] is not None:
            s = _advance(client, s["id"])
    return s


def _pdf_text(body: bytes) -> str:
    """Extract the text layer, recomposing SARA AM back to its composed form
    (the PDF stores it decomposed so glyphs map 1:1 to codepoints)."""
    import fitz
    with fitz.open(stream=body, filetype="pdf") as doc:
        raw = "\n".join(page.get_text() for page in doc)
    return raw.replace("ํา", "ำ")


def test_step_pdf_carries_the_methodology_appendix(client):
    s = _finished_consult(client)
    method = {"text": METHOD_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=method), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/report/positions.pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"

    text = _pdf_text(r.content)
    assert "หลักการและตรรกะที่ใช้ในรอบนี้" in text
    assert "Unit economics" in text and "Payback period" in text
    assert "ข้อมูลตัวเลขที่ถูกอ้างถึง" in text and "14" in text
    assert "กราฟเปรียบเทียบ" in text
    assert "ควรขยายสาขาไหม" in text          # the CEO's question
    assert "คำตอบเต็มของรอบนี้" in text       # the round itself, verbatim


def test_thai_text_layer_is_clean(client):
    """ที่ = ท + ◌ี + ◌่ — without harfbuzz shaping the tone mark collides with
    the vowel and is lost; ำ needs pre-decomposing or it copies out corrupted."""
    s = _finished_consult(client)
    method = {"text": METHOD_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=method), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/report/positions.pdf")
    text = _pdf_text(r.content)
    # ที่ = ◌ี + ◌่ · นี้ = ◌ี + ◌้ · คำ = the SARA AM that needs pre-decomposing
    for word in ("ที่ปรึกษา", "ที่มา", "คำตอบเต็มของรอบนี้", "หลักการและตรรกะที่ใช้ในรอบนี้"):
        assert word in text, f"lost in the text layer: {word}"
    # no substitution/control bytes leaking from unmapped glyphs
    assert not [c for c in text if ord(c) < 32 and c not in "\n\r\t"]


def test_methodology_is_cached_so_a_second_pdf_is_free(client):
    s = _finished_consult(client)
    method = {"text": METHOD_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=method) as m, \
         patch("app.llm.provider_ready", return_value=True):
        client.get(f"/api/consult/{s['id']}/report/positions.pdf")
        assert m.call_count == 1
        client.get(f"/api/consult/{s['id']}/report/positions.pdf")
        assert m.call_count == 1                       # served from the cache
        client.get(f"/api/consult/{s['id']}/report/positions.pdf?refresh=true")
        assert m.call_count == 2                       # explicit re-audit


def test_pdf_still_renders_when_the_audit_pass_fails(client):
    """No key, or non-JSON back: print the round, never invent an analysis."""
    s = _finished_consult(client)
    junk = {"text": "ขอโทษครับ ผมไม่เข้าใจ", "provider": "mock", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=junk):
        r = client.get(f"/api/consult/{s['id']}/report/positions.pdf")
    assert r.status_code == 200 and r.content[:4] == b"%PDF"
    text = _pdf_text(r.content)
    assert "ถอดวิธีคิดอัตโนมัติไม่สำเร็จ" in text
    assert "คำตอบเต็มของรอบนี้" in text


def test_executive_summary_offers_choices_and_a_recommendation(client):
    s = _finished_consult(client)
    opts = {"text": OPTIONS_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=opts), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/executive-summary.pdf")
    assert r.status_code == 200 and r.content[:4] == b"%PDF"

    text = _pdf_text(r.content)
    assert "บทสรุปสำหรับผู้บริหาร" in text
    # per-C-level takeaways, framed by the lane each is expert in
    assert "ข้อสรุปจากความเชี่ยวชาญของที่ปรึกษาแต่ละคน" in text
    assert "runway ไม่พอ" in text
    # the choices themselves + the board's pick + what only the CEO can settle
    assert "ทางเลือกในการตัดสินใจ" in text
    assert "นำร่อง 4 จุด" in text and "ยังไม่ขยาย" in text
    assert "ทางที่บอร์ดแนะนำ" in text
    assert "ประเด็นที่ CEO ต้องชี้ขาดเอง" in text
    assert "ช่องบันทึกการตัดสินใจของ CEO" in text


def test_options_endpoint_feeds_the_one_click_choices(client):
    s = _finished_consult(client)
    opts = {"text": OPTIONS_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=opts), \
         patch("app.llm.provider_ready", return_value=True):
        j = client.get(f"/api/consult/{s['id']}/options").json()
    assert j["recommended"] == "นำร่อง 4 จุด"
    assert [o["name"] for o in j["options"]] == ["นำร่อง 4 จุด", "ยังไม่ขยาย"]
    # cached onto the session so the UI can render them without another call
    assert client.get(f"/api/consults/{s['id']}").json()["options"]["recommended"] == "นำร่อง 4 จุด"


def test_executive_summary_before_the_brief_is_rejected(client):
    s = _framed(client)
    r = client.get(f"/api/consult/{s['id']}/executive-summary.pdf")
    assert r.status_code == 400 and "Stage 6" in r.text


def test_report_bad_inputs(client):
    s = _finished_consult(client)
    assert client.get(f"/api/consult/{s['id']}/report/nope.pdf").status_code == 400
    assert client.get(f"/api/consult/{s['id']}/report/research.pdf").status_code == 400  # never ran
    assert client.get("/api/consult/999/report/positions.pdf").status_code == 404
    assert client.get("/api/consult/999/executive-summary.pdf").status_code == 404
    assert client.get("/api/consult/999/options").status_code == 404


def test_chart_is_skipped_when_numbers_are_not_comparable(client):
    from app import report
    pdf = report._Doc("t")
    pdf.add_page()
    assert report._bar_chart(pdf, {"labels": ["a"], "values": [1]}) is False       # one bar
    assert report._bar_chart(pdf, {"labels": ["a", "b"], "values": [1]}) is False  # mismatched
    assert report._bar_chart(pdf, {"labels": ["a", "b"], "values": ["x", 2]}) is False
    assert report._bar_chart(pdf, {"labels": ["a", "b"], "values": [1, 2]}) is True


@pytest.mark.parametrize("raw", [
    '{"a": 1}',
    '```json\n{"a": 1}\n```',
    'นี่คือผลลัพธ์ครับ\n{"a": 1}\nหวังว่าจะช่วยได้',
])
def test_json_extraction_survives_chatty_models(client, raw):
    from app import report
    assert report._parse_json(raw) == {"a": 1}


@pytest.mark.parametrize("raw", ["", "ไม่มี JSON เลย", "{ยังไม่ปิดวงเล็บ", "[1,2,3]"])
def test_json_extraction_fails_closed(client, raw):
    from app import report
    assert report._parse_json(raw) is None


# ── Stage 5: red team, confidence ──

def test_red_team_attacks_the_framing_and_every_seat_scores_itself(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        s = _advance(client, s["id"])            # positions
        s = _advance(client, s["id"])            # debate

    def scored(provider, system, user, cancel=None, **kw):
        return _reply(f"จุดยืนหลังถกเถียง: คงเดิม\nความมั่นใจ: 62%", provider)

    with patch("app.llm.chat", side_effect=scored) as m:
        s = _advance(client, s["id"])
    results = s["steps"][-1]["results"]

    # the red team is briefed on what the debate cannot see: who was NOT invited
    # and how many genuinely different models are behind the seats
    red = [c for c in m.call_args_list if "Red Team อิสระ" in c.args[1]]
    assert len(red) == 1
    assert "ที่นั่งที่ไม่ถูกเรียก" in red[0].args[2]
    assert "จำนวน provider ที่ต่างกันจริง" in red[0].args[2]
    assert results["redteam"]["ok"] and "diversity" in results["redteam"]

    # every seat scores its own confidence, and the board average follows
    assert all(results[d]["confidence"] == 62 for d in config.DEPTS)
    assert s["confidence"]["average"] == 62
    assert s["confidence"]["lowest"] in config.DEPTS


@pytest.mark.parametrize(("text", "expected"), [
    ("ความมั่นใจ: 70%", 70),
    ("ความมั่นใจ 0%", 0),
    ("ความมั่นใจ: 100", 100),
    ("ความมั่นใจ: 120%", None),      # out of range, not a score
    ("มั่นใจมาก", None),
    ("", None),
])
def test_confidence_is_read_from_the_seat_reply(client, text, expected):
    from app import depts
    assert depts.parse_confidence(text) == expected


def test_the_brief_carries_confidence_into_the_pdf(client):
    s = _framed(client)

    def scored(provider, system, user, cancel=None, **kw):
        if "Defensible Brief" in system:
            return _reply("ข้อเสนอ: ทำแบบมีเงื่อนไข\nความมั่นใจ: 55%", provider)
        return _reply("จุดยืนหลังถกเถียง: คงเดิม\nความมั่นใจ: 55%", provider)

    with patch("app.llm.chat", side_effect=scored):
        while s["next_step"] is not None:
            s = _advance(client, s["id"])
    assert s["steps"][-1]["results"]["chair"]["confidence"] == 55

    with patch("app.llm.chat", side_effect=scored):
        r = client.get(f"/api/consult/{s['id']}/executive-summary.pdf")
    assert r.status_code == 200
    text = _pdf_text(r.content)
    assert "ความมั่นใจของบอร์ด: 55%" in text
    assert "ความมั่นใจที่แต่ละที่นั่งให้ตัวเอง" in text
    assert "Red Team" in text                     # dissent and its cause stay on the record


# ── echo-chamber control: distinct models behind the seats ──

def test_the_default_line_up_is_diverse_but_not_enforced(client):
    """The shipped default puts every seat on a different agent — personas on one
    model is one brain in five hats. But it is a *default*, not a rule: the CEO
    may put two seats on one lab deliberately, and the board's job is to warn,
    never to refuse. This test guards both halves of that."""
    assigned = set(config.DEFAULT_PROVIDERS.values())
    assert len(assigned) == len(config.DEFAULT_PROVIDERS), "the default must not share an agent"
    assert len(assigned) >= 4

    # any seat may take any provider, including one already in use elsewhere
    for dept in config.DEPTS:
        for provider in config.PROVIDERS:
            r = client.put(f"/api/dept/{dept}/provider", json={"provider": provider})
            assert r.status_code == 200, f"{dept} -> {provider} was refused: {r.text}"
            assert r.json()["providers"][dept] == provider

    # and the extreme case — one lab everywhere — is allowed but called out
    for dept in config.DEPTS:
        client.put(f"/api/dept/{dept}/provider", json={"provider": "anthropic"})
    state = client.get("/api/state").json()
    assert state["diversity"]["distinct"] == 1
    assert state["diversity"]["warning"], "a single-lab board must be flagged"


def test_agents_can_be_reset_to_the_shipped_line_up(client):
    """Free choice needs an undo: after shuffling five seats there must be a way
    back to the diverse default without hand-editing JSON."""
    for dept in config.DEPTS:
        client.put(f"/api/dept/{dept}/provider", json={"provider": "mock"})
    assert set(client.get("/api/state").json()["depts"][0].values())  # sanity

    r = client.post("/api/agents/reset")
    assert r.status_code == 200
    assert r.json()["providers"] == {d: config.DEFAULT_PROVIDERS.get(d, "mock")
                                     for d in config.DEPTS}
    assert r.json()["diversity"]["distinct"] >= 1


def test_reset_touches_agents_only_not_the_board_s_history(client):
    """Resetting who answers must not erase what the board already concluded."""
    s = _finished_consult(client)
    before = len(client.get("/api/consults").json()["consults"])
    client.post("/api/agents/reset")
    after = client.get("/api/consults").json()["consults"]
    assert len(after) == before
    assert any(c["id"] == s["id"] for c in after), "the consult survived the reset"


def test_an_unknown_seat_or_agent_is_still_refused(client):
    """Free choice among real agents — not free choice of typos, which would
    silently leave a seat on a provider that does not exist."""
    assert client.put("/api/dept/nope/provider", json={"provider": "anthropic"}).status_code == 404
    assert client.put("/api/dept/cmo/provider", json={"provider": "gpt-9"}).status_code == 400


def test_a_single_vendor_board_is_called_out(client, monkeypatch):
    """Five seats on one lab is one brain in five hats — the board must say so
    rather than let unanimous agreement look like corroboration."""
    from app import depts
    monkeypatch.setattr("app.llm.provider_ready", lambda p: p == "zai")
    for dept in config.DEPTS:
        client.put(f"/api/dept/{dept}/provider", json={"provider": "zai"})
    div = depts.model_diversity()
    assert div["distinct"] == 1
    assert div["warning"] and "Z.AI" in div["warning"]
    assert client.get("/api/state").json()["diversity"]["distinct"] == 1


def test_two_models_from_one_lab_count_as_one_vendor(client, monkeypatch):
    """Opus and Fable are two provider keys but one lab: same corpus, same
    refusal habits, same blind spots. Counting keys would have reported a
    diverse board while two seats shared a brain."""
    from app import depts
    monkeypatch.setattr("app.llm.provider_ready",
                        lambda p: p in ("anthropic", "anthropic_fable"))
    client.put("/api/dept/cfo/provider", json={"provider": "anthropic_fable"})
    client.put("/api/dept/researcher/provider", json={"provider": "anthropic"})
    for dept in ("cmo", "coo", "datalyst"):
        client.put(f"/api/dept/{dept}/provider", json={"provider": "mock"})

    div = depts.model_diversity()
    assert div["distinct"] == 1, "two Anthropic models are one vendor"
    assert div["live"] == 2, "both seats are live"
    assert div["shared_vendors"].get("Anthropic") == ["cfo", "researcher"]
    assert div["warning"] and "Anthropic" in div["warning"]


def test_vendor_overlap_is_named_seat_by_seat(client, monkeypatch):
    """The CEO needs to know *which* seats share a lab, not just that some do."""
    from app import depts
    monkeypatch.setattr("app.llm.provider_ready", lambda p: p != "mock")
    assignment = {"cmo": "gemini", "cfo": "anthropic_fable", "coo": "zai",
                  "researcher": "anthropic", "datalyst": "deepseek"}
    for dept, prov in assignment.items():
        client.put(f"/api/dept/{dept}/provider", json={"provider": prov})

    div = depts.model_diversity()
    assert div["distinct"] == 4, "Anthropic twice -> four labs across five seats"
    assert set(div["per_vendor"]) == {"Anthropic", "Google", "Z.AI", "DeepSeek"}
    assert sorted(div["shared_vendors"]["Anthropic"]) == ["cfo", "researcher"]
    # the warning must name the seats, using their display names
    assert "CFO" in div["warning"] and "Researcher" in div["warning"]


def test_every_provider_declares_a_vendor(client):
    """A provider without a vendor label would silently become its own lab and
    inflate the diversity count."""
    from app import depts
    for key, meta in config.PROVIDERS.items():
        assert meta.get("vendor"), f"{key} declares no vendor"
        assert depts.vendor_of(key) == meta["vendor"]
    # an unknown key must not crash the count; it stands in as its own lab
    assert depts.vendor_of("does-not-exist") == "does-not-exist"


def test_a_fully_diverse_board_raises_no_warning(client, monkeypatch):
    from app import depts
    monkeypatch.setattr("app.llm.provider_ready", lambda p: p != "mock")
    for dept, prov in {"cmo": "gemini", "cfo": "manus", "coo": "zai",
                       "researcher": "anthropic", "datalyst": "deepseek"}.items():
        client.put(f"/api/dept/{dept}/provider", json={"provider": prov})
    div = depts.model_diversity()
    assert div["distinct"] == 5 and div["shared_vendors"] == {}
    assert div["warning"] is None


def test_no_live_provider_is_called_out(client, monkeypatch):
    from app import depts
    monkeypatch.setattr("app.llm.provider_ready", lambda p: False)
    div = depts.model_diversity()
    assert div["live"] == 0 and "mock" in div["warning"]


# ── persistent memory across sessions ──

MEMORY_JSON = json.dumps({
    "conclusion": "ชะลอการขยายสาขาจนกว่า runway จะเกิน 12 เดือน",
    "stance": "ทำแบบมีเงื่อนไข", "confidence": 70,
    "constraints": ["ห้ามผูกสัญญาเช่าเกิน 1 ปีจนกว่าจะพิสูจน์ยอดต่อสาขา"],
    "open_questions": ["ยอดต่อสาขาที่แท้จริง"],
    "tripwires": ["runway ต่ำกว่า 6 เดือน"],
}, ensure_ascii=False)


def test_a_finished_session_is_filed_to_memory(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        for _ in range(3):
            s = _advance(client, s["id"])
    with patch("app.llm.chat", return_value=_reply(MEMORY_JSON)):
        s = _advance(client, s["id"])             # brief -> distil -> file

    saved = client.get("/api/memory").json()["memory"]
    assert len(saved) == 1
    assert saved[0]["conclusion"].startswith("ชะลอการขยายสาขา")
    assert saved[0]["consult_id"] == s["id"] and saved[0]["confidence"] == 70
    assert client.get("/api/state").json()["memory_count"] == 1


def test_a_failed_distil_does_not_break_the_brief(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        while s["next_step"] is not None:
            s = _advance(client, s["id"])
    assert s["status"] == "done"                  # brief still landed
    assert client.get("/api/memory").json()["memory"] == []


def test_frame_flags_a_question_that_cuts_against_a_past_ruling(client):
    from app import store
    store.add_memory({"consult_id": 1, "question": "ควรขยายสาขาไหม",
                      "project": None, "conclusion": "ชะลอการขยาย",
                      "stance": "ไม่ทำ", "confidence": 80,
                      "constraints": ["ห้ามผูกสัญญาเกิน 1 ปี"],
                      "open_questions": [], "tripwires": []})
    conflict = json.dumps({
        "conflicts": [{"memory_id": 1, "past": "เคยสรุปว่าให้ชะลอการขยาย",
                       "tension": "คำถามนี้กลับมาขยายอีกครั้งโดยยังไม่มีหลักฐานใหม่",
                       "severity": "สูง"}],
        "carry_forward": ["ห้ามผูกสัญญาเกิน 1 ปี"],
    }, ensure_ascii=False)

    def reply(provider, system, user, cancel=None, **kw):
        if "decision consistency auditor" in system:
            assert "ควรขยายสาขาไหม" in user      # the archive is what it audits against
            return _reply(conflict, provider)
        return _reply(_frame_json(), provider)

    s = _start(client)
    with patch("app.llm.chat", side_effect=reply):
        s = _advance(client, s["id"])
    framer = s["steps"][0]["results"]["framer"]
    assert framer["memory_checked"] == 1
    assert framer["conflicts"][0]["memory_id"] == 1
    assert framer["carry_forward"] == ["ห้ามผูกสัญญาเกิน 1 ปี"]

    # and the constraint the board already committed to reaches the seats
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    assert all("ห้ามผูกสัญญาเกิน 1 ปี" in c.args[2] for c in m.call_args_list)


def test_a_hallucinated_memory_reference_is_dropped(client):
    from app import memory, store
    store.add_memory({"consult_id": 1, "question": "q", "project": None,
                      "conclusion": "c", "stance": "ทำ", "confidence": 50,
                      "constraints": [], "open_questions": [], "tripwires": []})
    bogus = json.dumps({"conflicts": [{"memory_id": 99, "past": "x", "tension": "y"}],
                        "carry_forward": []}, ensure_ascii=False)
    with patch("app.llm.chat", return_value=_reply(bogus)):
        out = memory.conflicts("คำถามใหม่", None, "anthropic")
    assert out["conflicts"] == [] and out["checked"] == 1


def test_memory_is_scoped_per_project_and_can_be_forgotten(client):
    from app import store
    a = store.add_memory({"question": "q1", "project": "YourFin", "conclusion": "c1"})
    store.add_memory({"question": "q2", "project": "FlowerVending", "conclusion": "c2"})
    assert [m["project"] for m in store.get_memory("YourFin")] == ["YourFin"]
    assert len(client.get("/api/memory").json()["memory"]) == 2
    assert len(client.get("/api/memory?project=YourFin").json()["memory"]) == 1
    assert client.delete(f"/api/memory/{a['id']}").status_code == 200
    assert client.delete(f"/api/memory/{a['id']}").status_code == 404


# ── asking a general question, with the business library left out ──

def test_a_general_question_leaves_the_document_library_out(client):
    client.post("/api/line/webhook", json={"events": [{"type": "message",
        "message": {"type": "text", "text": "ธุรกิจของฉันคือตู้กดดอกไม้ที่เอกมัย"}}]})

    r = client.post("/api/consult", json={"question": "SaaS pricing ควรคิดยังไง",
                                          "project": "__none__", "web_research": False})
    s = r.json()
    assert s["use_docs"] is False and s["project"] is None

    with patch("app.llm.chat", return_value=_reply(_frame_json())) as m:
        s = _advance(client, s["id"])
    assert all("ตู้กดดอกไม้" not in c.args[2] for c in m.call_args_list)

    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    for c in m.call_args_list:
        assert "คลังเอกสารธุรกิจของ CEO" not in c.args[2]
        assert "ตู้กดดอกไม้" not in c.args[2]


def test_the_library_is_still_used_when_no_project_is_picked(client):
    """Empty project means "every project", which is not the same as "none"."""
    client.post("/api/line/webhook", json={"events": [{"type": "message",
        "message": {"type": "text", "text": "ธุรกิจของฉันคือตู้กดดอกไม้ที่เอกมัย"}}]})
    s = _framed(client, question="ควรขยายไหม")
    assert s["use_docs"] is True
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    assert all("ตู้กดดอกไม้" in c.args[2] for c in m.call_args_list)


def test_a_doc_free_session_keeps_that_choice_when_branched(client):
    r = client.post("/api/consult", json={"question": "q", "project": "__none__",
                                          "web_research": False})
    s = r.json()
    with patch("app.llm.chat", return_value=_reply(_frame_json())):
        s = _advance(client, s["id"])
    child = client.post(f"/api/consult/{s['id']}/branch", json={"step": "frame"}).json()
    assert child["use_docs"] is False


# ── decisions: reopen the question, or erase what it taught the board ──

def _decided(client):
    """A finished consult, its memory filed, and a decision recorded against it."""
    s = _framed(client, question="ควรขยายสาขาไหม")
    with patch("app.llm.chat", side_effect=_fake_chat):
        for _ in range(3):
            s = _advance(client, s["id"])
    with patch("app.llm.chat", return_value=_reply(MEMORY_JSON)):
        s = _advance(client, s["id"])
    d = client.post("/api/decisions", json={"consult_id": s["id"],
                                            "question": s["question"],
                                            "decision": "ขยาย 3 สาขา"}).json()
    return s, d


def test_rethink_reopens_the_question_without_touching_the_original(client):
    s, d = _decided(client)
    r = client.post(f"/api/decisions/{d['id']}/rethink",
                    json={"direction": "คราวนี้ให้มองมุมชะลอการลงทุน"})
    assert r.status_code == 200
    child = r.json()
    assert child["id"] != s["id"] and child["parent_id"] == s["id"]
    assert child["branched_from"] == "decision"
    assert child["question"] == s["question"] and child["steps"] == []
    assert child["next_step"] == "frame"          # a fresh framing, not a resumed stage
    assert child["direction"] == "คราวนี้ให้มองมุมชะลอการลงทุน"

    original = client.get(f"/api/consults/{s['id']}").json()
    assert original["status"] == "done" and len(original["steps"]) == 5


def test_rethink_inherits_the_scope_of_the_consult_it_came_from(client):
    r = client.post("/api/consult", json={"question": "q", "project": "__none__",
                                          "web_research": False})
    s = r.json()
    with patch("app.llm.chat", return_value=_reply(_frame_json())):
        s = _advance(client, s["id"])
    d = client.post("/api/decisions", json={"consult_id": s["id"], "question": "q",
                                            "decision": "ลุย"}).json()
    child = client.post(f"/api/decisions/{d['id']}/rethink", json={}).json()
    assert child["use_docs"] is False and child["web_research"] is False


def test_forgetting_a_decision_stops_it_steering_later_sessions(client):
    s, d = _decided(client)
    assert len(client.get("/api/memory").json()["memory"]) == 1

    out = client.post(f"/api/decisions/{d['id']}/forget").json()
    assert out["forgotten"] == 1 and out["consult_id"] == s["id"]
    assert client.get("/api/memory").json()["memory"] == []
    # the decision itself survives — only the board's learning was erased
    assert len(client.get("/api/decisions").json()["decisions"]) == 1

    # and the next session is no longer audited against it
    with patch("app.llm.chat", return_value=_reply(_frame_json())) as m:
        s2 = _start(client, question="ควรขยายอีกไหม")
        _advance(client, s2["id"])
    assert not [c for c in m.call_args_list if "decision consistency auditor" in c.args[1]]


def test_a_decision_with_no_consult_has_nothing_to_forget(client):
    d = client.post("/api/decisions", json={"question": "โจทย์", "decision": "ลุย"}).json()
    out = client.post(f"/api/decisions/{d['id']}/forget").json()
    assert out["forgotten"] == 0 and out["reason"]


def test_deleting_a_decision_can_take_its_learning_with_it(client):
    s, d = _decided(client)
    out = client.delete(f"/api/decisions/{d['id']}?forget=true").json()
    assert out["deleted"] == d["id"] and out["forgotten"] == 1
    assert client.get("/api/decisions").json()["decisions"] == []
    assert client.get("/api/memory").json()["memory"] == []


def test_deleting_a_decision_can_keep_its_learning(client):
    s, d = _decided(client)
    assert client.delete(f"/api/decisions/{d['id']}").json()["forgotten"] == 0
    assert len(client.get("/api/memory").json()["memory"]) == 1


def test_decision_actions_404_on_unknown_ids(client):
    assert client.post("/api/decisions/999/rethink", json={}).status_code == 404
    assert client.post("/api/decisions/999/forget").status_code == 404
    assert client.delete("/api/decisions/999").status_code == 404


# ── documents: removing what went stale ──

def test_deleting_a_document_removes_it_from_the_board_and_from_disk(client):
    import app.docs as docs_mod
    client.post("/api/docs/projects", json={"name": "YourFin"})
    up = client.post("/api/docs/upload",
                     files={"file": ("old-rent.txt", "ค่าเช่าเดิม 85,000".encode(), "text/plain")},
                     data={"project": "YourFin"}).json()
    path = docs_mod.LOCAL_DIR / "YourFin" / "old-rent.txt"
    assert path.exists()
    assert "ค่าเช่าเดิม" in docs_mod.knowledge_context()

    out = client.delete(f"/api/docs/{up['id']}").json()
    assert out["deleted"] == up["id"] and out["local"] is True and out["errors"] == []
    assert not path.exists()
    assert client.get("/api/docs").json()["documents"] == []
    # and the board stops quoting it on the next consult
    assert "ค่าเช่าเดิม" not in docs_mod.knowledge_context()


def test_deleting_a_document_survives_a_missing_file(client):
    """The metadata entry must go even when the file is already gone, or the
    library keeps listing a document nobody can open."""
    import app.docs as docs_mod
    up = client.post("/api/docs/upload",
                     files={"file": ("gone.txt", b"x", "text/plain")}).json()
    (docs_mod.LOCAL_DIR / "gone.txt").unlink()
    out = client.delete(f"/api/docs/{up['id']}").json()
    assert out["deleted"] == up["id"] and out["local"] is False
    assert client.get("/api/docs").json()["documents"] == []


def test_deleting_an_unknown_document_404s(client):
    assert client.delete("/api/docs/999").status_code == 404


# ── archived sessions from the pre-Crucible pipeline ──

def test_old_sessions_stay_readable_and_are_never_resumed(client):
    from app import store
    s = store.create_session("โจทย์เก่า")
    store.append_step(s["id"], "opinions", {"cfo": {"text": "x", "provider": "m", "ok": True}})
    store.append_step(s["id"], "synthesis", {"chair": {"text": "y", "provider": "m", "ok": True}})

    view = client.get(f"/api/consults/{s['id']}").json()
    assert view["legacy"] is True and view["next_step"] is None
    assert view["step_labels"]["opinions"].endswith("(รูปแบบเดิม)")
    # and its executive summary still exports, off the old key
    with patch("app.llm.chat", side_effect=_fake_chat):
        assert client.get(f"/api/consult/{s['id']}/executive-summary.pdf").status_code == 200


# ── guardrails / prompt integrity ──

def test_guardrails_in_prompts(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    position_systems = [c.args[1] for c in m.call_args_list if "กฎเหล็ก" in c.args[1]]
    assert len(position_systems) == len(config.DEPTS)
    joined = "\n".join(position_systems)
    assert "ห้ามออกความเห็นเรื่องสภาพคล่อง" in joined      # CMO guardrail
    assert "ห้ามออกความเห็นเรื่องกลยุทธ์การตลาด" in joined  # CFO guardrail
    for sys_prompt in position_systems:
        for part in ("จุดยืน", "หลักฐานที่หนุนจุดยืนนี้", "สิ่งที่จะทำให้ผมเปลี่ยนใจ"):
            assert part in sys_prompt


def test_the_debate_attacks_evidence_not_opinions(client):
    s = _framed(client, web=True)
    with patch("app.llm.chat", side_effect=_fake_chat), \
         patch("app.research.gather", return_value={"sources": FAKE_SOURCES, "errors": []}):
        s = _advance(client, s["id"])            # research
        s = _advance(client, s["id"])            # positions
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])                # debate
    for c in m.call_args_list:
        assert "จงโจมตี **หลักฐาน** ของเขา" in c.args[1]
        assert "จุดยืนของที่ปรึกษาคนอื่น" in c.args[2]
        # a seat can only attack evidence it can see, so the others' sources travel with it
        assert "หลักฐานที่คนอื่นใช้" in c.args[2]


def test_the_brief_reads_the_whole_session(client):
    s = _framed(client)
    with patch("app.llm.chat", side_effect=_fake_chat):
        for _ in range(3):                       # positions, debate, redteam
            s = _advance(client, s["id"])
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        s = _advance(client, s["id"])
    # one chair writing the brief, plus one pass that files it to memory
    brief_calls = [c for c in m.call_args_list if "Defensible Brief" in c.args[1]]
    assert len(brief_calls) == 1
    system, user = brief_calls[0].args[1], brief_calls[0].args[2]
    assert "ห้ามกลบเสียงค้าน" in system and "ความมั่นใจ" in system
    assert "Stage 1" in user and "Stage 5" in user


# ── decisions ──

def test_decision_log_and_scoring(client):
    r = client.post("/api/decisions", json={"consult_id": 1, "question": "โจทย์", "decision": "ลุย"})
    d = r.json()
    assert d["id"] == 1 and d["verdict"] is None
    r = client.post("/api/decisions/1/score", json={"outcome": "รอด", "verdict": "saved"})
    assert r.json()["verdict"] == "saved"
    j = client.get("/api/decisions").json()
    assert j["stats"]["total"] == 1 and j["stats"]["saved"] == 1
    assert client.post("/api/decisions/1/score", json={"outcome": "", "verdict": "nope"}).status_code == 400
    assert client.post("/api/decisions/99/score", json={"outcome": "", "verdict": "saved"}).status_code == 404


def test_every_provider_accepts_the_cancel_probe(client):
    """chat() hands `cancel` to whichever caller runs — a provider added without
    that parameter would blow up mid-consult instead of at import time.

    The first three params are the contract every caller must honour; optional
    extras after them (e.g. max_tokens) are allowed but must be keyword-safe
    with a default, since chat() only passes them to opted-in providers.
    """
    import inspect

    from app import llm
    for name, caller in llm._CALLERS.items():
        params = inspect.signature(caller).parameters
        assert list(params)[:3] == ["system", "user", "cancel"], f"{name} signature drifted"
        assert params["cancel"].default is None, f"{name} must default cancel to None"
        for extra in list(params)[3:]:
            assert params[extra].default is not inspect.Parameter.empty, \
                f"{name}.{extra} must be optional — chat() does not always pass it"
    # every provider promising max_tokens support must actually accept it
    for name in llm._ACCEPTS_MAX_TOKENS:
        assert "max_tokens" in inspect.signature(llm._CALLERS[name]).parameters, \
            f"{name} is listed in _ACCEPTS_MAX_TOKENS but takes no max_tokens"
    # a provider missing from any of the three tables is a half-wired provider
    from app import config
    assert set(llm._CALLERS) == set(llm._HAS_KEY) == set(config.PROVIDERS)
    assert llm._ACCEPTS_MAX_TOKENS <= set(llm._CALLERS)


def test_provider_switch(client):
    r = client.put("/api/dept/cfo/provider", json={"provider": "anthropic"})
    assert r.json()["providers"]["cfo"] == "anthropic"
    assert client.put("/api/dept/cfo/provider", json={"provider": "nope"}).status_code == 400
    assert client.put("/api/dept/nope/provider", json={"provider": "mock"}).status_code == 404


def test_index_serves_advisory_ui(client):
    html = client.get("/").text
    for marker in ("view-board", "view-decisions", "view-docs", "view-agents",
                   "askBoard", "recordDecision", "setProvider",
                   "loadDocs", "syncDrive", "step_labels",
                   # new core-flow controls
                   "ask-project", "uploadFiles", "stopBoard", "resetTo", "branchAt",
                   "seatResearchCard", "chairCard", "ask-web", "gateBlock", "dot",
                   # Crucible stages
                   "frameCard", "redteamCard", "researchBlock", "convened",
                   # Pipeline: one work tree per routine
                   "view-pipeline", "loadPipeline", "createRoutine", "commentTask",
                   "branchTask", "thinkPanel", "rt-owner", "rt-project"):
        assert marker in html, marker
    # `runTask` used to be listed here, back when the hub deliberately executed
    # nothing. The Pipeline page executes recurring work on the CEO's button, so
    # the marker is now real; what must stay gone is the old per-department
    # automation UI this hub replaced.
    for stale in ("view-cmo", "LLM Learning"):
        assert stale not in html, f"stale automation marker: {stale}"


# ── documents ──

def test_docs_line_webhook_saves_text(client):
    payload = {"events": [{"type": "message",
                           "message": {"type": "text", "text": "ยอดขายเดือนนี้ 120,000 บาท"}}]}
    r = client.post("/api/line/webhook", json=payload)
    assert r.status_code == 200 and r.json()["saved"] == 1
    j = client.get("/api/docs").json()
    assert j["documents"][0]["source"] == "line"
    assert "120,000" in j["documents"][0]["text"]
    assert j["knowledge_chars"] > 0
    assert j["drive_connected"] is False  # no SA json in test env -> local mirror


def test_upload_from_browser(client):
    r = client.post("/api/docs/upload",
                    files={"file": ("plan.txt", b"runway 8 months", "text/plain")},
                    data={"project": ""})
    assert r.status_code == 200
    doc = r.json()
    assert doc["source"] == "upload" and doc["name"] == "plan.txt"
    assert "runway 8 months" in doc["text"]
    assert client.get("/api/docs").json()["documents"][0]["name"] == "plan.txt"


def test_upload_into_a_chosen_project_skips_the_librarian(client):
    client.post("/api/docs/projects", json={"name": "YourFin"})
    with patch("app.docs.classify_project") as guess:
        r = client.post("/api/docs/upload",
                        files={"file": ("loan.csv", b"a,b\n1,2", "text/csv")},
                        data={"project": "YourFin"})
    assert r.json()["project"] == "YourFin"
    guess.assert_not_called()   # the CEO already said where it goes
    import app.docs as docs
    assert (docs.LOCAL_DIR / "YourFin" / "loan.csv").exists()


def test_empty_upload_is_rejected(client):
    r = client.post("/api/docs/upload", files={"file": ("x.txt", b"", "text/plain")})
    assert r.status_code == 400


def test_docs_knowledge_feeds_consult(client):
    client.post("/api/line/webhook", json={"events": [{"type": "message",
        "message": {"type": "text", "text": "ธุรกิจของฉันคือตู้กดดอกไม้ที่เอกมัย"}}]})
    s = _framed(client, question="ควรขยายไหม")
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    users = [c.args[2] for c in m.call_args_list]
    assert all("คลังเอกสารธุรกิจของ CEO" in u and "ตู้กดดอกไม้" in u for u in users)


def test_consult_scoped_to_one_project_ignores_other_projects(client):
    client.post("/api/docs/projects", json={"name": "YourFin"})
    client.post("/api/docs/projects", json={"name": "FlowerVending"})
    with patch("app.llm.chat", return_value={"text": "YourFin", "provider": "gemini",
                                             "model": "m", "ok": True}), \
         patch("app.llm.provider_ready", return_value=True):
        client.post("/api/line/webhook", json={"events": [{"type": "message",
            "message": {"type": "text", "text": "ตารางผ่อนมือถือ งวดละ 1,200 บาท"}}]})
    with patch("app.llm.chat", return_value={"text": "FlowerVending", "provider": "gemini",
                                             "model": "m", "ok": True}), \
         patch("app.llm.provider_ready", return_value=True):
        client.post("/api/line/webhook", json={"events": [{"type": "message",
            "message": {"type": "text", "text": "ตู้ดอกไม้เอกมัยยอด 40,000"}}]})

    s = _start(client, question="ควรขยายไหม", project="YourFin")
    with patch("app.llm.chat", side_effect=_fake_chat) as m:
        _advance(client, s["id"])
    user = m.call_args.args[2]
    assert "ผ่อนมือถือ" in user and "ตู้ดอกไม้เอกมัย" not in user


def test_docs_sync_without_drive_returns_error(client):
    j = client.post("/api/docs/sync").json()
    assert j["synced"] == 0 and "Drive" in j["error"]


def test_librarian_projects_and_classification(client):
    r = client.post("/api/docs/projects", json={"name": "YourFin"})
    assert "YourFin" in r.json()["projects"]
    client.post("/api/docs/projects", json={"name": "FlowerVending"})
    assert client.post("/api/docs/projects", json={"name": "a/b"}).status_code == 400

    with patch("app.llm.chat", return_value={"text": "YourFin", "provider": "gemini",
                                             "model": "m", "ok": True}), \
         patch("app.llm.provider_ready", return_value=True):
        client.post("/api/line/webhook", json={"events": [{"type": "message",
            "message": {"type": "text", "text": "ตารางผ่อนมือถือ งวดละ 1,200 บาท"}}]})
    j = client.get("/api/docs").json()
    doc = j["documents"][0]
    assert doc["project"] == "YourFin"
    import app.docs as docs
    assert (docs.LOCAL_DIR / "YourFin" / doc["name"]).exists()
    assert "### โปรเจค: YourFin" in docs.knowledge_context()


def test_librarian_reclassify_unfiled(client):
    with patch("app.llm.chat", side_effect=_fake_chat):  # returns non-project text -> unfiled
        client.post("/api/line/webhook", json={"events": [{"type": "message",
            "message": {"type": "text", "text": "โน้ตยังไม่เข้าโปรเจค"}}]})
    client.post("/api/docs/projects", json={"name": "YourFin"})
    with patch("app.llm.chat", return_value={"text": "YourFin", "provider": "gemini",
                                             "model": "m", "ok": True}), \
         patch("app.llm.provider_ready", return_value=True):
        j = client.post("/api/docs/reclassify").json()
    assert j["filed"] == 1 and j["unfiled"] == 0


# ── CFO financial model (Excel scenario forecast) ──

FINMODEL_JSON = json.dumps({
    "business_model": "ขายช่อดอกไม้ผ่านตู้กดอัตโนมัติ",
    "currency_note": "บาท (THB), 24 เดือน",
    "revenue": {
        "units_month_1": {"value": 300, "source": "CMO: ขายได้ 300 ช่อเดือนแรก"},
        "price_per_unit": {"value": 150, "source": "บทถกเถียง: ช่อละ 150 บาท"},
        "growth_rate_monthly": {"value": 0.04, "source": "ประมาณการของ CFO"},
        "churn_or_seasonality": {"value": 0.01, "source": "ประมาณการของ CFO"},
    },
    "costs": {
        "cogs_pct_of_revenue": {"value": 0.40, "source": "CFO: ต้นทุนดอก 40%"},
        "fixed_cost_month": {"value": 25000, "source": "COO: ค่าเช่า+คนดูแล"},
        "marketing_pct_of_revenue": {"value": 0.08, "source": "ประมาณการของ CFO"},
        "other_variable_pct": {"value": 0.05, "source": "ประมาณการของ CFO"},
    },
    "capital": {
        "upfront_investment": {"value": 400000, "source": "COO: ค่าตู้ 4 แสน"},
        "starting_cash": {"value": 600000, "source": "CFO: เงินสดในมือ"},
        "payment_collection_lag_months": {"value": 0, "source": "ขายเงินสด"},
    },
    "scenarios": {
        "base": {"revenue_mult": 1.0, "cost_mult": 1.0, "rationale": "กรณีฐาน"},
        "best": {"revenue_mult": 1.3, "cost_mult": 0.95, "rationale": "ทำเลดีกว่าคาด"},
        "worst": {"revenue_mult": 0.65, "cost_mult": 1.2, "rationale": "ดอกเน่าเสียสูง"},
    },
    "risks": [{"risk": "ดอกไม้เน่าเสียเกิน 15%", "driver": "cogs_pct_of_revenue",
               "trigger": "ของเสียเกิน 15% สองเดือนติด", "mitigation": "ลดสต็อกต่อรอบเติม"}],
    "kpis_to_watch": ["ยอดขายต่อตู้ต่อวัน", "อัตราของเสีย", "เงินสดคงเหลือ"],
    "sensitivity": {"driver_a": "revenue.growth_rate_monthly",
                    "driver_b": "costs.cogs_pct_of_revenue",
                    "why": "สองตัวนี้ชี้ว่าคุ้มทุนเมื่อไหร่"},
}, ensure_ascii=False)


def _finmodel_wb(client, payload=FINMODEL_JSON):
    """Run a consult to completion, then fetch the workbook openpyxl-loaded."""
    from openpyxl import load_workbook
    s = _finished_consult(client)
    fake = {"text": payload, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=fake), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/financial-model.xlsx")
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK", "xlsx must be a zip container"
    return s, r, load_workbook(io.BytesIO(r.content))


def test_financial_model_has_every_scenario_sheet(client):
    _s, r, wb = _finmodel_wb(client)
    assert "spreadsheetml.sheet" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    for name in ("วิธีอ่าน", "สมมติฐาน", "Base", "Best", "Worst",
                 "เปรียบเทียบ Scenario", "ความเสี่ยง"):
        assert name in wb.sheetnames, f"missing sheet {name}"


def test_assumptions_are_the_only_typed_numbers(client):
    """The CEO must be able to change one input and trust everything follows —
    so no scenario cell may hold a hardcoded *number*. Text headers are fine;
    a literal value is not, because it would silently ignore the assumptions.
    """
    _s, _r, wb = _finmodel_wb(client)
    for label in ("Base", "Best", "Worst"):
        ws = wb[label]
        for row in ws.iter_rows(min_row=6, min_col=2):
            for cell in row:
                assert not isinstance(cell.value, (int, float)), (
                    f"{label}!{cell.coordinate} holds the literal {cell.value!r} — "
                    "it must be a formula referencing 'สมมติฐาน'"
                )


def test_scenario_formulas_reference_the_assumptions_sheet(client):
    _s, _r, wb = _finmodel_wb(client)
    ws = wb["Base"]
    joined = " ".join(str(c.value) for row in ws.iter_rows(min_row=6, min_col=2)
                      for c in row if c.value)
    assert "สมมติฐาน" in joined, "scenario sheet must point back at the assumptions"


def test_estimates_are_labelled_apart_from_debate_numbers(client):
    """A number the CFO invented must never look like one the board argued."""
    _s, _r, wb = _finmodel_wb(client)
    ws = wb["สมมติฐาน"]
    kinds = {}
    for row in ws.iter_rows(min_row=7, max_col=5):
        label, kind = row[0].value, row[4].value
        if label and kind:
            kinds[label] = kind
    assert kinds["อัตราเติบโตต่อเดือน"] == "ประมาณการ"
    assert kinds["ต้นทุนขาย (% ของรายได้)"] == "จากบทถกเถียง"


def _evaluate(xlsx_bytes: bytes):
    """Compute the workbook with a real formula engine and return a cell reader.

    `formulas` keys cells as "'[filename]SHEETNAME'!REF" — the sheet name is
    upper-cased but the filename keeps its original case, so build the key the
    same way rather than guessing.
    """
    formulas = pytest.importorskip("formulas")
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "model.xlsx")
        with open(path, "wb") as fh:
            fh.write(xlsx_bytes)
        sol = formulas.ExcelModel().loads(path).finish().calculate()
    prefix = f"'[{os.path.basename(path)}]"

    def cell(sheet: str, ref: str):
        return sol[f"{prefix}{sheet.upper()}'!{ref}"].value[0, 0]
    return cell


def _row_of(ws, label: str) -> int:
    """Find a P&L line by its Thai label instead of hardcoding a row number —
    inserting a line into the model must not silently retarget the assertion."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == label:
            return row[0].row
    raise AssertionError(f"row {label!r} not found in {ws.title}")


def test_workbook_math_is_correct_when_evaluated(client):
    """Formulas are only useful if they compute the right thing — evaluate the
    real workbook and check the P&L against an independent Python model."""
    _s, r, wb = _finmodel_wb(client)
    cell = _evaluate(r.content)
    ws = wb["Base"]
    rev_r = _row_of(ws, "รายได้")
    ni_r = _row_of(ws, "กำไรสุทธิ (Net Profit)")
    out_r = _row_of(ws, "เงินสดจ่าย")
    cum_r = _row_of(ws, "เงินสดคงเหลือสะสม")

    # independent recomputation of month 1 and month 2 of the Base case
    units1, price, growth, churn = 300, 150, 0.04, 0.01
    cogs_pct, fixed, mkt, oth = 0.40, 25000, 0.08, 0.05
    var_pct = cogs_pct + mkt + oth
    rev1 = units1 * price
    ni1 = rev1 * (1 - var_pct) - fixed
    units2 = units1 * (1 + growth - churn)
    ni2 = units2 * price * (1 - var_pct) - fixed

    assert cell("Base", f"B{rev_r}") == pytest.approx(rev1, rel=1e-6)
    assert cell("Base", f"B{ni_r}") == pytest.approx(ni1, rel=1e-6)
    assert cell("Base", f"C{ni_r}") == pytest.approx(ni2, rel=1e-6)
    # cash paid out is every variable cost plus fixed cost, carried as negative
    assert cell("Base", f"B{out_r}") == pytest.approx(-(rev1 * var_pct + fixed), rel=1e-6)
    # cash on hand must absorb the upfront investment, so it trails profit
    assert cell("Base", f"B{cum_r}") == pytest.approx(600000 + ni1 - 400000, rel=1e-6)
    # month 2 has no capex, so cash grows by that month's net cash flow
    assert cell("Base", f"C{cum_r}") == pytest.approx(
        600000 + ni1 - 400000 + ni2, rel=1e-6)


def test_scenario_multipliers_actually_move_the_numbers(client):
    """Best/Worst must diverge from Base, otherwise the scenarios are theatre."""
    _s, r, wb = _finmodel_wb(client)
    cell = _evaluate(r.content)
    rev_r = _row_of(wb["Base"], "รายได้")
    base, best, worst = (cell(s, f"B{rev_r}") for s in ("Base", "Best", "Worst"))
    assert best > base > worst, f"revenue must rank best>base>worst (got {best}/{base}/{worst})"
    assert best == pytest.approx(base * 1.3, rel=1e-6)
    assert worst == pytest.approx(base * 0.65, rel=1e-6)


def test_financial_model_refuses_before_the_board_has_spoken(client):
    """No debate means no assumptions — better a clear 400 than invented numbers."""
    s = _start(client)
    with patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/financial-model.xlsx")
    assert r.status_code == 400
    assert "สมมติฐาน" in r.json()["detail"]


def test_financial_model_survives_an_unusable_cfo_reply(client):
    """A truncated or non-JSON reply must 400, never a half-built workbook."""
    s = _finished_consult(client)
    junk = {"text": '{"revenue": {"units_month_1"', "provider": "anthropic",
            "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=junk), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/financial-model.xlsx")
    assert r.status_code == 400


def test_read_timeout_scales_with_max_tokens(client):
    """Raising max_tokens without raising the read timeout turns a slow-but-fine
    reply into a failure — this is the bug that broke the 8192-token Thai
    financial model against the flat 120s budget."""
    from app import llm

    assert llm._timeout_for(None) == llm.TIMEOUT
    assert llm._timeout_for(2048) == llm.TIMEOUT      # default budget unchanged
    assert llm._timeout_for(8192) > llm.TIMEOUT       # the model's real ask
    # monotonic: more tokens must never mean less time
    budgets = [llm._timeout_for(n) for n in (2048, 4096, 8192, 16384)]
    assert budgets == sorted(budgets)

    seen = {}

    class _Resp:
        status_code = 200
        def raise_for_status(self): pass
        def json(self): return {"content": [{"type": "text", "text": "ok"}]}

    def _capture(url, **kw):
        seen["timeout"] = kw.get("timeout")
        return _Resp()

    with patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.httpx.post", side_effect=_capture):
        llm._anthropic("s", "u", None, max_tokens=8192)
        big = seen["timeout"]
        llm._anthropic("s", "u")
        small = seen["timeout"]
    assert big > small, "the 8192-token call must get a longer read timeout"


def test_financial_assumptions_endpoint_exposes_provenance(client):
    s = _finished_consult(client)
    fake = {"text": FINMODEL_JSON, "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=fake), \
         patch("app.llm.provider_ready", return_value=True):
        j = client.get(f"/api/consult/{s['id']}/financial-assumptions").json()
    assert j["revenue"]["price_per_unit"]["value"] == 150
    assert "ประมาณการ" in j["revenue"]["growth_rate_monthly"]["source"]
    assert client.get("/api/consult/999/financial-assumptions").status_code == 404


def test_dirty_assumption_values_do_not_break_the_model(client):
    """Providers return '1,250 บาท' and '3%' — coerce instead of crashing."""
    messy = json.loads(FINMODEL_JSON)
    messy["revenue"]["price_per_unit"] = {"value": "1,250 บาท", "source": "x"}
    messy["revenue"]["growth_rate_monthly"] = {"value": "3%", "source": "x"}
    messy["costs"]["fixed_cost_month"] = {"value": None, "source": "x"}
    _s, _r, wb = _finmodel_wb(client, json.dumps(messy, ensure_ascii=False))
    ws = wb["สมมติฐาน"]
    vals = {row[0].value: row[1].value for row in ws.iter_rows(min_row=7, max_col=2)}
    assert vals["ราคาขายต่อหน่วย (บาท)"] == 1250
    assert vals["อัตราเติบโตต่อเดือน"] == pytest.approx(0.03)
    assert vals["ค่าใช้จ่ายคงที่ต่อเดือน (บาท)"] == 0


# ── departmental deliverables (documents + peer review) ──

# Derived from SPECS, so a department added there is covered by every
# parametrized deliverable test without touching this file.
_DOC_DEPTS = set(deliverable.SPECS)


def _deliverable_json(dept: str) -> str:
    """A plausible reply for whichever spec is asked for, built from the spec
    itself so a section added to SPECS can't silently go untested."""
    from app import deliverable as D
    sections = {}
    for key, label, kind, _guide in D.SPECS[dept]["sections"]:
        if kind == "prose":
            sections[key] = {"body": f"เนื้อหา {label} อ้างอิง [1]"}
        elif kind == "bullets":
            sections[key] = {"items": [f"งาน {label} ข้อ 1", "งาน ข้อ 2 (ประมาณการ)"]}
        else:
            sections[key] = {"columns": ["ประเด็น", "ค่า", "แหล่ง [n]"],
                             "rows": [[f"{label} แถว 1", "1,200", "[1]"],
                                      [f"{label} แถว 2", "800", "(ประมาณการ)"]]}
            if kind == "matrix":
                sections[key]["chart"] = {"title": label, "labels": ["ก", "ข"],
                                          "values": [1200, 800], "unit": "บาท"}
    return json.dumps({
        "headline": f"ข้อสรุปเอกสาร {dept}",
        "sections": sections,
        "confidence": "กลาง — ข้อมูลคู่แข่งยังบาง",
        "data_gaps": ["ยอดขายจริงรายสาขา", "ราคาคู่แข่งในทำเลเดียวกัน"],
    }, ensure_ascii=False)


def _deliverable(client, dept, payload=None):
    """Finish a consult, then fetch a department's document."""
    s = _finished_consult(client)
    fake = {"text": payload or _deliverable_json(dept), "provider": "anthropic",
            "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=fake), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/deliverable/{dept}.pdf")
    return s, r


def test_every_seat_hands_over_an_artefact(client):
    """A seat with no deliverable is a seat that debates and produces nothing.
    CFO's artefact is the Excel model, so it is exempt from the PDF specs."""
    from app import deliverable as D

    covered = set(D.SPECS) | {"cfo"}          # cfo -> finmodel.py
    assert covered >= set(config.DEPTS), f"seats with no artefact: {set(config.DEPTS) - covered}"
    assert "cfo" not in D.SPECS, "the finance artefact is the workbook, not a PDF"
    for dept, spec in D.SPECS.items():
        assert spec["dept"] == dept, f"{dept} spec is mislabelled as {spec['dept']}"
        keys = [k for k, *_ in spec["sections"]]
        assert len(keys) == len(set(keys)), f"{dept} has duplicate section keys"
        assert keys[-1] == "actions", f"{dept} must end with what to do next"
        assert all(kind in D._KIND_SHAPE for _k, _l, kind, _g in spec["sections"])


def test_frontend_offers_a_button_for_every_artefact(client):
    """SPECS and the UI's DELIVERABLES map must not drift apart, or a document
    exists on the server that the CEO has no way to open."""
    from app import deliverable as D

    html = (Path(__file__).resolve().parent.parent / "static" / "index.html").read_text("utf-8")
    block = html.split("const DELIVERABLES = {", 1)[1].split("};", 1)[0]
    for dept in set(D.SPECS) | {"cfo"}:
        assert f"{dept}:" in block, f"no download button wired for {dept}"


@pytest.mark.parametrize("dept", sorted(_DOC_DEPTS))
def test_deliverable_covers_every_section_the_ceo_asked_for(client, dept):
    """Each section title in the spec must actually reach the paper — a spec
    entry that never renders is a promise the CEO can't see."""
    from app import deliverable as D
    _s, r = _deliverable(client, dept)
    assert r.status_code == 200, r.text
    assert r.content[:4] == b"%PDF"
    text = _pdf_text(r.content)
    for _key, label, _kind, _guide in D.SPECS[dept]["sections"]:
        assert label in text, f"{dept} deliverable is missing section {label!r}"


def test_cmo_deliverable_carries_the_marketing_disciplines(client):
    _s, r = _deliverable(client, "cmo")
    text = _pdf_text(r.content)
    for heading in ("สำรวจตลาด", "วิเคราะห์คู่แข่ง", "จุดยืนแบรนด์",
                    "โอกาสทางตลาด", "ตัวชี้วัดหลัก", "ประมาณการยอดขาย"):
        assert heading in text


def test_coo_deliverable_carries_the_operations_disciplines(client):
    _s, r = _deliverable(client, "coo")
    text = _pdf_text(r.content)
    for heading in ("วินิจฉัยกระบวนการ", "Core Flow", "Systemization",
                    "Lean", "Quality & Risk", "Performance Measurement"):
        assert heading in text


@pytest.mark.parametrize("dept", sorted(_DOC_DEPTS))
def test_deliverable_is_peer_reviewed_before_it_reaches_the_ceo(client, dept):
    """The collaboration requirement: the other three advisors critique the
    document and the author answers, all on the same page as the work."""
    _s, r = _deliverable(client, dept)
    text = _pdf_text(r.content)
    assert "บอร์ดวิพากษ์เอกสารนี้" in text
    others = [d for d in config.DEPTS if d != dept]
    for o in others:
        name = config.DEPTS[o]["name"]
        assert name in text, f"{name} did not review the {dept} document"
    assert "คำชี้แจงของ" in text, "the author never answered the critique"


@pytest.mark.parametrize("dept", sorted(_DOC_DEPTS))
def test_deliverable_reviewers_exclude_the_author(client, dept):
    """An advisor grading its own homework would defeat the review."""
    from app import deliverable as D
    s = _finished_consult(client)
    fake = {"text": _deliverable_json(dept), "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=fake), \
         patch("app.llm.provider_ready", return_value=True):
        data = D.build(s, dept)
    assert dept not in data["review"]["critiques"]
    assert set(data["review"]["critiques"]) == set(config.DEPTS) - {dept}


def test_deliverable_cites_the_web_sources_it_was_given(client):
    """Claims must trace to a real URL, so the source table has to be printed."""
    s = _start(client, question="ควรขยายสาขาไหม", web=True)
    with patch("app.llm.chat", side_effect=_fake_chat), \
         patch("app.research.search_detail",
               return_value={"results": FAKE_SOURCES, "error": None, "engine": "test"}), \
         patch("app.research.fetch_text", return_value="เนื้อหาเต็มของหน้า"):
        for _ in range(5):
            if s["next_step"] is None:
                break
            s = _advance(client, s["id"])
    fake = {"text": _deliverable_json("cmo"), "provider": "anthropic", "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=fake), \
         patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/deliverable/cmo.pdf")
    assert r.status_code == 200
    text = _pdf_text(r.content)
    assert "แหล่งอ้างอิงที่บอร์ดมีให้ใช้" in text
    assert "example.com/market" in text, "the cited URL must be printed for the CEO"


def test_deliverable_prompt_forbids_faking_citations_without_research(client):
    """With no web round, the author must be told not to invent [n] markers."""
    from app import deliverable as D
    s = _finished_consult(client)          # web_research off
    ground = D._grounding(s)
    assert "ห้ามใส่ [n]" in ground
    assert "(ประมาณการ)" in ground


def test_deliverable_author_prompt_demands_provenance(client):
    from app import deliverable as D
    sys_prompt = D._author_system(D.SPECS["cmo"])
    assert "ห้ามใส่ [n] ปลอม" in sys_prompt
    assert "ห้ามแต่งชื่อคู่แข่ง" in sys_prompt
    # every section of the spec must be described to the model
    for key, _label, _kind, _guide in D.SPECS["cmo"]["sections"]:
        assert f'"{key}"' in sys_prompt


def test_deliverable_refuses_unusable_replies_and_unknown_depts(client):
    s = _finished_consult(client)
    junk = {"text": "ขอโทษครับ ผมทำเอกสารนี้ไม่ได้", "provider": "anthropic",
            "model": "m", "ok": True}
    with patch("app.llm.chat", return_value=junk), \
         patch("app.llm.provider_ready", return_value=True):
        assert client.get(f"/api/consult/{s['id']}/deliverable/cmo.pdf").status_code == 400
    # CFO's artefact is the workbook, not a PDF, so this route must not invent one
    assert client.get(f"/api/consult/{s['id']}/deliverable/cfo.pdf").status_code == 404
    assert client.get(f"/api/consult/{s['id']}/deliverable/nope.pdf").status_code == 404
    assert client.get("/api/consult/999/deliverable/cmo.pdf").status_code == 404


def test_deliverable_needs_a_debate_first(client):
    s = _start(client)
    with patch("app.llm.provider_ready", return_value=True):
        r = client.get(f"/api/consult/{s['id']}/deliverable/cmo.pdf")
    assert r.status_code == 400
    assert "บทถกเถียง" in r.json()["detail"]


def test_deliverable_survives_ragged_tables(client):
    """Providers return short rows and stray columns — pad, don't crash."""
    payload = json.dumps({
        "headline": "ทดสอบตารางไม่สมบูรณ์",
        "sections": {
            "market_survey": {"columns": ["a", "b", "c"], "rows": [["1"], ["1", "2", "3", "4"]]},
            "competitors": {"columns": [], "rows": []},
            "brand_position": {"body": ""},
            "opportunity": {"columns": ["x"], "rows": [["y"]],
                            "chart": {"labels": ["only-one"], "values": [1]}},
            "core_metrics": "ไม่ใช่ dict เลย",
            "sales_forecast": {"columns": ["q"], "rows": "ไม่ใช่ list"},
            "actions": {"items": []},
        },
    }, ensure_ascii=False)
    _s, r = _deliverable(client, "cmo", payload)
    assert r.status_code == 200 and r.content[:4] == b"%PDF"
    text = _pdf_text(r.content)
    assert "ไม่พบข้อมูลสำหรับหัวข้อนี้" in text


def test_deliverable_is_cached_then_rebuilt_on_refresh(client):
    """Re-downloading must not re-bill the CEO for the same document."""
    from app import deliverable as D
    from app import store
    s = _finished_consult(client)
    fake = {"text": _deliverable_json("coo"), "provider": "anthropic", "model": "m", "ok": True}
    def reload():
        got = store.get_consult(s["id"])
        assert got is not None
        return got

    with patch("app.llm.chat", return_value=fake) as chat, \
         patch("app.llm.provider_ready", return_value=True):
        D.build(s, "coo")
        first = chat.call_count
        D.build(reload(), "coo")                            # cached
        assert chat.call_count == first
        D.build(reload(), "coo", refresh=True)
        assert chat.call_count > first


# ── board seats & agent assignment ──

def test_every_seat_has_a_lane_a_default_agent_and_a_working_caller(client):
    """A seat added to config.DEPTS must be wired everywhere at once — a seat
    with no lane speaks unguarded, and one with no provider silently mocks."""
    from app import depts, llm

    for dept in config.DEPTS:
        assert dept in depts.LANES, f"{dept} has no guardrail lane"
        lane, guard = depts.LANES[dept]
        assert lane and guard, f"{dept}'s lane is empty"
        provider = config.DEFAULT_PROVIDERS.get(dept)
        assert provider, f"{dept} has no default agent"
        assert provider in config.PROVIDERS, f"{dept} points at unknown agent {provider}"
        assert provider in llm._CALLERS, f"{provider} has no caller"
        assert provider in llm._HAS_KEY, f"{provider} has no key probe"


def test_the_researcher_seat_exists_and_is_evidence_only(client):
    """The Researcher reports what the evidence says; it must not hand out
    marketing/finance/ops strategy — that is what the other seats are for."""
    from app import depts

    assert "researcher" in config.DEPTS
    s = client.get("/api/state").json()
    seat = next(d for d in s["depts"] if d["key"] == "researcher")
    assert seat["name"] == "Researcher"
    lane, guard = depts.LANES["researcher"]
    assert "หลักฐาน" in lane
    assert "ห้าม" in guard and "(ไม่มีหลักฐาน)" in guard


def test_prompts_enumerate_seats_from_config_not_a_frozen_list(client):
    """The methodology/options prompts used to hand-list four departments, so a
    new seat was dropped from every report without any test failing."""
    from app import report

    for prompt in (report.METHOD_SYSTEM, report.OPTIONS_SYSTEM):
        for dept in config.DEPTS:
            assert dept in prompt, f"{dept} missing from a board prompt"
    assert "researcher" in report.OPTIONS_SYSTEM


def test_frontend_derives_seats_from_the_api(client):
    """The board grid read a hardcoded DEPT_KEYS array, which would hide any
    seat the server added."""
    html = (Path(__file__).resolve().parent.parent / "static" / "index.html").read_text("utf-8")
    assert "DEPT_KEYS" not in html, "frontend still hardcodes the seat list"
    assert "deptKeys()" in html
    assert "repeat(4, 1fr)" not in html, "round grid still assumes exactly 4 seats"


def test_a_new_seat_is_added_to_an_existing_store(client, tmp_path):
    """A store written before the Researeer existed must gain the seat on load,
    while the CEO's own overrides stay untouched."""
    import json

    from app import store

    legacy = {"providers": {"cmo": "manus"}, "consults": [], "decisions": []}
    store._FILE.write_text(json.dumps(legacy), encoding="utf-8")
    got = store.get_providers()
    assert got["cmo"] == "manus", "an explicit CEO choice must never be overwritten"
    assert got["researcher"] == config.DEFAULT_PROVIDERS["researcher"]
    assert set(got) >= set(config.DEPTS)


def test_each_seat_runs_the_agent_the_ceo_assigned(client):
    """The exact line-up the CEO specified, asserted as model ids so a silent
    provider rename or a stale .env pin cannot pass."""
    expected = {
        "coo": "glm-5.2",
        "cmo": "gemini-3.1-pro-preview",
        "cfo": "claude-fable-5",
        "researcher": "claude-sonnet-5",
        "datalyst": "deepseek-v4-pro",
    }
    for dept, model in expected.items():
        provider = config.DEFAULT_PROVIDERS[dept]
        actual = config.PROVIDERS[provider]["model"]
        assert actual == model, f"{dept} should run {model}, got {actual}"


def test_fable_and_opus_are_separate_seats_on_one_key(client):
    """Fable and Opus share the Anthropic key but must stay distinct choices."""
    from app import llm

    assert config.PROVIDERS["anthropic"]["model"] != config.PROVIDERS["anthropic_fable"]["model"]
    sent = {}

    class _R:
        status_code = 200
        def raise_for_status(self): pass
        def json(self): return {"content": [{"type": "text", "text": "ok"}]}

    with patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.httpx.post", side_effect=lambda url, **kw: (sent.update(kw["json"]), _R())[1]):
        llm._anthropic_fable("s", "u")
        assert sent["model"] == config.PROVIDERS["anthropic_fable"]["model"]
        llm._anthropic("s", "u")
        assert sent["model"] == config.PROVIDERS["anthropic"]["model"]


def test_deepseek_speaks_openai_shape(client):
    from app import llm

    sent = {}

    class _R:
        status_code = 200
        def raise_for_status(self): pass
        def json(self): return {"choices": [{"message": {"content": " ok "}}]}

    with patch("app.config.DEEPSEEK_API_KEY", "k"), \
         patch("app.llm.httpx.post", side_effect=lambda url, **kw: (sent.update(kw["json"]), _R())[1]):
        assert llm._deepseek("sys", "usr", None, max_tokens=4096) == "ok"
    assert sent["model"] == config.PROVIDERS["deepseek"]["model"]
    assert [m["role"] for m in sent["messages"]] == ["system", "user"]
    assert sent["max_tokens"] == 4096


def test_every_seat_with_a_url_has_a_service_directory(client):
    """A seat whose URL points at nothing shows a permanent offline dot — that
    was the Researcher's state until its evidence desk existed."""
    services = Path(__file__).resolve().parent.parent.parent / "services"
    for dept in config.DEPTS:
        d = services / dept
        assert d.is_dir(), f"{dept} has no service directory ({d})"
        assert any(d.glob("*.py")) or any(d.glob("*.js")) or (d / "src").is_dir(), \
            f"{dept}'s service directory has no server"


def test_the_launcher_starts_every_department_service(client):
    """start_all.sh listed five ports while the board had six services, so the
    readiness check passed with the evidence desk still down."""
    sh = (Path(__file__).resolve().parent.parent.parent / "scripts" / "start_all.sh").read_text("utf-8")
    for port in ("8100", "8201", "8202", "8203", "8204", "8205"):
        assert port in sh, f"port {port} is never started"
    assert "researcher" in sh, "the evidence desk is never spawned"
    # the wait loop and the report must read one shared list, not two literals
    assert sh.count("for p in $PORTS") == 2
    assert "-sTCP:LISTEN" in sh, "port_busy would count a CLOSE_WAIT socket as a server"


def test_truncated_json_is_salvaged_not_discarded():
    """A Thai document can blow the output budget mid-write. Losing the tail is
    acceptable; losing every finished section is not."""
    from app import report

    full = {"headline": "ข้อสรุป",
            "sections": {"a": {"body": "เนื้อหา ก"}, "b": {"body": "เนื้อหา ข"}}}
    intact = json.dumps(full, ensure_ascii=False)
    assert report._parse_json(intact) == full          # untruncated path unchanged

    cut = intact[:intact.index('"b"') + 24]            # chopped inside section b
    got = report._parse_json(cut)
    assert got is not None, "salvage returned nothing"
    assert got["headline"] == "ข้อสรุป"
    assert "a" in got["sections"] and got["sections"]["a"] == {"body": "เนื้อหา ก"}
    assert "b" not in got["sections"], "a half-written section must be dropped, not guessed"


def test_salvage_refuses_garbage_rather_than_inventing_structure():
    from app import report

    assert report._parse_json("") is None
    assert report._parse_json("ขอโทษครับ ทำไม่ได้") is None
    assert report._parse_json("{") is None
    assert report._parse_json('{"a"') is None
    # a truncated string containing braces must not fool the brace counter
    assert report._parse_json('{"a": "ข้อความมี { และ } อยู่ข้างใน') is None


def test_salvaged_documents_admit_what_is_missing(client):
    """A short document must not read as 'nothing more to report'."""
    from app import deliverable as D

    dept = "cmo"
    spec = D.SPECS[dept]
    first_key, first_label = spec["sections"][0][0], spec["sections"][0][1]
    partial = json.dumps({"headline": "บางส่วน",
                          "sections": {first_key: {"columns": ["a"], "rows": [["1"]]}}},
                         ensure_ascii=False)
    _s, r = _deliverable(client, dept, payload=partial)
    assert r.status_code == 200
    text = _pdf_text(r.content)
    assert "เอกสารนี้ไม่สมบูรณ์" in text
    assert first_label not in text.split("เอกสารนี้ไม่สมบูรณ์")[1], \
        "a section that WAS written must not be listed as missing"
    last_label = spec["sections"][-1][1]
    assert last_label in text, "the unwritten section must be named"


def test_transient_provider_failures_are_retried(client):
    """Anthropic answering 529 'overloaded' once must not cost a deliverable
    that took a minute of board debate to earn."""
    from app import llm

    calls = {"n": 0}

    def flaky(system, user, cancel=None, max_tokens=None):
        calls["n"] += 1
        if calls["n"] < 3:
            raise httpx.HTTPStatusError(
                "overloaded", request=httpx.Request("POST", "https://x"),
                response=httpx.Response(529, request=httpx.Request("POST", "https://x")))
        return "recovered"

    with patch.dict(llm._CALLERS, {"anthropic": flaky}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.time.sleep"):                     # no real backoff in tests
        out = llm.chat("anthropic", "s", "u")
    assert out["ok"] and out["text"] == "recovered"
    assert calls["n"] == 3


def test_permanent_failures_are_not_retried(client):
    """A 401 will fail identically every time — burning three attempts on it
    just makes the CEO wait longer for the same error."""
    from app import llm

    calls = {"n": 0}

    def unauthorized(system, user, cancel=None, max_tokens=None):
        calls["n"] += 1
        raise httpx.HTTPStatusError(
            "bad key", request=httpx.Request("POST", "https://x"),
            response=httpx.Response(401, request=httpx.Request("POST", "https://x")))

    with patch.dict(llm._CALLERS, {"anthropic": unauthorized}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.time.sleep"):
        out = llm.chat("anthropic", "s", "u")
    assert not out["ok"]
    assert calls["n"] == 1, "a 401 must fail fast"


def test_a_stop_beats_a_retry(client):
    """The CEO pressing STOP must not be made to sit through the backoff."""
    from app import llm

    calls = {"n": 0}

    def always_529(system, user, cancel=None, max_tokens=None):
        calls["n"] += 1
        raise httpx.HTTPStatusError(
            "overloaded", request=httpx.Request("POST", "https://x"),
            response=httpx.Response(529, request=httpx.Request("POST", "https://x")))

    with patch.dict(llm._CALLERS, {"anthropic": always_529}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.time.sleep"):
        out = llm.chat("anthropic", "s", "u", cancel=lambda: calls["n"] >= 1)
    assert not out["ok"]
    assert calls["n"] == 1, "cancel must be honoured between attempts"


# ── history / bad inputs ──

def test_consult_history_and_delete(client):
    s = _start(client)
    assert client.get("/api/consults").json()["consults"][0]["id"] == s["id"]
    assert client.delete(f"/api/consults/{s['id']}").status_code == 200
    assert client.get(f"/api/consults/{s['id']}").status_code == 404
    assert client.delete(f"/api/consults/{s['id']}").status_code == 404


def test_bad_inputs(client):
    assert client.post("/api/consult", json={"question": "  "}).status_code == 400
    assert client.post("/api/decisions", json={"question": "", "decision": ""}).status_code == 400
    assert client.get("/api/consults/999").status_code == 404
    assert client.post("/api/consult/999/advance", json={}).status_code == 404
    assert client.post("/api/consult/999/reset", json={"step": "opinions"}).status_code == 404


# ── reading what the AI actually sent back ──
#
# Everything below pins one class of bug: the provider answered, the hub decided
# it understood, and the CEO was shown something the model never said. A silent
# misread is more expensive than an error, so each of these has to fail loudly.


@pytest.mark.parametrize(("raw", "expected"), [
    ('{"a": 1}', {"a": 1}),
    ("```json\n{\"a\": 1}\n```", {"a": 1}),                     # fenced
    ("```\n{\"a\": 1}\n```", {"a": 1}),                          # fenced, unlabelled
    ('นี่คือผลลัพธ์ครับ\n{"a": 1}\nหวังว่าจะช่วยได้', {"a": 1}),   # wrapped in prose
    ('สรุป (ตามที่ขอ) {ดังนี้} ครับ\n{"a": 1}', {"a": 1}),        # a brace in the preamble
    ('{"a": 1,}', {"a": 1}),                                     # trailing comma
    ('{"a": 1, /* หมายเหตุ */ "b": 2}', {"a": 1, "b": 2}),        # block comment
    ('{"a": 1} // เสร็จแล้ว', {"a": 1}),                          # line comment
    ('{"url": "https://x.com/a"}', {"url": "https://x.com/a"}),  # ...that is not a URL
    ('{"a": True, "b": None}', {"a": True, "b": None}),          # python literals
    ('{"a": "บรรทัดแรก\nบรรทัดสอง"}', {"a": "บรรทัดแรก\nบรรทัดสอง"}),  # raw newline in a string
    ('{“a”: 1}', {"a": 1}),                                      # typographic quotes
])
def test_every_way_a_model_mangles_its_json(client, raw, expected):
    """Each of these is a real provider habit. `find("{")..rfind("}")` — what the
    three old copies of _parse_json did — gets the brace-in-the-preamble case
    wrong and every repair case wrong."""
    from app import jsonx
    assert jsonx.extract(raw) == expected


@pytest.mark.parametrize("raw", ["", "   ", "ไม่มี JSON เลย", "{ยังไม่ปิดวงเล็บ", "[1,2,3]", None])
def test_unreadable_json_fails_closed(client, raw):
    """None, never a plausible object nobody wrote."""
    from app import jsonx
    assert jsonx.extract(raw) is None


def test_a_truncated_reply_keeps_the_sections_that_finished(client):
    """A reply cut at the token ceiling used to cost the whole document. Keep the
    entries that closed, and prefer the outermost object over some nested
    fragment that happens to parse on its own."""
    from app import jsonx
    cut = '{"sections": {"a": "หนึ่ง", "b": "สอง"}, "risk": "ครึ่งประ'
    assert jsonx.extract(cut) == {"sections": {"a": "หนึ่ง", "b": "สอง"}}


@pytest.mark.parametrize(("raw", "unit", "value"), [
    ("1,250 บาท", None, 1250),          # thousands separator + currency word
    ("1.2 ล้านบาท", None, 1_200_000),   # Thai magnitude, silently dropped before
    ("4,200 ล้าน", None, 4.2e9),
    ("50k", None, 50_000),
    ("(3,000)", None, -3000),            # accounting negative, read as +3000 before
    ("100-200", None, 150),              # a range is not a subtraction
    ("3 เดือน", None, 3),                # "m" for months must not mean millions
    ("3%", "pct", 0.03),
    (3, "pct", 0.03),                    # the schema said 0.03; the model said 3
    (0.03, "pct", 0.03),                 # …and an already-correct rate is left alone
    (40, "pct", 0.4),
    (125, "mult", 1.25),                 # a scenario multiplier written as a percent
    (1.25, "mult", 1.25),
    ("ไม่ทราบ", None, 0),
    (None, None, 0),
    (True, None, 0),                     # a bool is not a number
])
def test_numbers_are_read_the_way_a_thai_cfo_writes_them(client, raw, unit, value):
    from app import jsonx
    assert jsonx.number(raw, unit) == pytest.approx(value)


def test_a_reinterpreted_unit_is_declared_not_silently_fixed(client):
    """The CEO must be able to trace every number. A unit the hub corrected on his
    behalf is exactly the number he would never think to check."""
    from app import jsonx
    value, note = jsonx.number_detail(3, "pct")
    assert value == pytest.approx(0.03) and "3%" in note
    assert jsonx.number_detail(0.03, "pct")[1] == "", "a correct value needs no note"


def test_percent_confusion_reaches_the_workbook_as_a_rate_and_a_warning(client):
    """`3` in a percent field is a 300%/month forecast — ten orders of magnitude
    of error by month 24, agreed with by every sheet, chart and break-even."""
    messy = json.loads(FINMODEL_JSON)
    messy["revenue"]["growth_rate_monthly"] = {"value": 3, "source": "บอร์ด"}
    messy["costs"]["fixed_cost_month"] = {"value": "1.2 ล้านบาท", "source": "บอร์ด"}
    messy["scenarios"]["best"]["revenue_mult"] = 125
    _s, _r, wb = _finmodel_wb(client, json.dumps(messy, ensure_ascii=False))
    ws = wb["สมมติฐาน"]
    rows = {row[0].value: row for row in ws.iter_rows(min_row=7, max_col=5)}
    assert rows["อัตราเติบโตต่อเดือน"][1].value == pytest.approx(0.03)
    assert "ระบบตีความหน่วย" in (rows["อัตราเติบโตต่อเดือน"][3].value or "")
    assert rows["ค่าใช้จ่ายคงที่ต่อเดือน (บาท)"][1].value == pytest.approx(1_200_000)
    # the scenario multipliers are read with the same eye
    best = [r for r in ws.iter_rows(min_row=7, max_col=3) if r[0].value == "Best"][0]
    assert best[1].value == pytest.approx(1.25)


# ── the transport layer: an answer that isn't there, and one that was cut ──

class _Resp:
    """Minimal httpx.Response stand-in for the provider transports."""

    status_code = 200

    def __init__(self, payload):
        self._payload = payload

    def raise_for_status(self):
        pass

    def json(self):
        return self._payload


def test_gemini_reads_the_answer_not_the_scratchpad(client):
    """Gemini 3 returns its reasoning as a part flagged `thought: true`, usually
    first. `parts[0].text` therefore hands the board the model's scratchpad, and
    drops any answer split across several parts."""
    from app import llm

    payload = {"candidates": [{"content": {"parts": [
        {"text": "ขอคิดก่อน...", "thought": True},
        {"text": "ส่วนที่หนึ่ง"},
        {"text": "ส่วนที่สอง"},
    ]}, "finishReason": "STOP"}]}
    with patch("app.config.GOOGLE_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(payload)):
        assert llm._gemini("s", "u") == "ส่วนที่หนึ่ง\nส่วนที่สอง"


def test_gemini_says_when_it_was_cut_or_blocked(client):
    from app import llm

    cut = {"candidates": [{"content": {"parts": [{"text": "ครึ่งเดียว"}]},
                           "finishReason": "MAX_TOKENS"}]}
    with patch("app.config.GOOGLE_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(cut)):
        with pytest.raises(llm.Truncated) as e:
            llm._gemini("s", "u")
        assert e.value.text == "ครึ่งเดียว"       # the partial survives the raise

    blocked = {"promptFeedback": {"blockReason": "SAFETY"}, "candidates": []}
    with patch("app.config.GOOGLE_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(blocked)):
        with pytest.raises(RuntimeError, match="SAFETY"):
            llm._gemini("s", "u")


def test_a_reply_that_is_all_thinking_is_a_failure_not_an_answer(client):
    """Claude returning only thinking blocks, or DeepSeek spending the budget in
    `reasoning_content`, used to reach the board as a confident blank section."""
    from app import llm

    thinking_only = {"content": [{"type": "thinking", "thinking": "..."}],
                     "stop_reason": "end_turn"}
    with patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(thinking_only)):
        with pytest.raises(llm.EmptyReply):
            llm._anthropic("s", "u")

    reasoning_only = {"choices": [{"message": {"content": "", "reasoning_content": "..."},
                                   "finish_reason": "stop"}]}
    with patch("app.config.DEEPSEEK_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(reasoning_only)):
        with pytest.raises(llm.EmptyReply):
            llm._deepseek("s", "u")


def test_every_provider_reports_the_token_ceiling_in_its_own_dialect(client):
    """stop_reason / finish_reason / finishReason — three spellings of the same
    event, none of which were being read."""
    from app import llm

    claude_cut = {"content": [{"type": "text", "text": "ครึ่ง"}], "stop_reason": "max_tokens"}
    with patch("app.config.ANTHROPIC_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(claude_cut)):
        with pytest.raises(llm.Truncated):
            llm._anthropic("s", "u", None, max_tokens=100)

    openai_cut = {"choices": [{"message": {"content": "ครึ่ง"}, "finish_reason": "length"}]}
    with patch("app.config.DEEPSEEK_API_KEY", "k"), \
         patch("app.llm.httpx.post", return_value=_Resp(openai_cut)):
        with pytest.raises(llm.Truncated):
            llm._deepseek("s", "u")


def test_chat_hands_back_the_partial_and_flags_the_cut(client):
    """A cut reply is still worth salvaging — but the caller has to be told, or a
    half-written document reads as a finished one."""
    from app import llm

    def cut(system, user, cancel=None, max_tokens=None):
        raise llm.Truncated("ครึ่งเดียว", "anthropic")

    with patch.dict(llm._CALLERS, {"anthropic": cut}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), patch("app.llm.time.sleep"):
        out = llm.chat("anthropic", "s", "u")
    assert out["truncated"] and out["ok"] and out["text"] == "ครึ่งเดียว"


def test_an_empty_reply_is_retried_but_a_cut_one_is_not(client):
    """An empty completion is often a bad minute; the same budget cuts at the same
    place every time, so retrying a truncation just burns the CEO's clock."""
    from app import llm

    calls = {"empty": 0, "cut": 0}

    def empty_then_fine(system, user, cancel=None, max_tokens=None):
        calls["empty"] += 1
        if calls["empty"] == 1:
            raise llm.EmptyReply("nothing")
        return "recovered"

    def always_cut(system, user, cancel=None, max_tokens=None):
        calls["cut"] += 1
        raise llm.Truncated("ครึ่ง", "anthropic")

    with patch.dict(llm._CALLERS, {"anthropic": empty_then_fine}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), patch("app.llm.time.sleep"):
        assert llm.chat("anthropic", "s", "u")["text"] == "recovered"
    assert calls["empty"] == 2

    with patch.dict(llm._CALLERS, {"anthropic": always_cut}), \
         patch("app.config.ANTHROPIC_API_KEY", "k"), patch("app.llm.time.sleep"):
        llm.chat("anthropic", "s", "u")
    assert calls["cut"] == 1


# ── chat_json: ask for a schema, come back with one ──

def _chat_reply(text, ok=True, truncated=False):
    return {"text": text, "provider": "anthropic", "model": "m",
            "ok": ok, "truncated": truncated}


def test_a_good_json_reply_costs_exactly_one_call(client):
    from app import llm
    with patch("app.llm.chat", return_value=_chat_reply('{"a": 1}')) as m:
        out = llm.chat_json("anthropic", "s", "u", required=("a",))
    assert out["data"] == {"a": 1} and out["calls"] == 1 and m.call_count == 1


def test_a_model_that_answers_in_prose_is_shown_what_broke(client):
    """Re-asking blind costs the same tokens for a worse hit rate — quote the
    broken reply back and the model fixes its own format."""
    from app import llm
    replies = [_chat_reply("ขอโทษครับ ผมไม่เข้าใจ"), _chat_reply('{"a": 1}')]
    with patch("app.llm.chat", side_effect=replies) as m:
        out = llm.chat_json("anthropic", "s", "u", required=("a",))
    assert out["data"] == {"a": 1} and out["calls"] == 2
    repair = m.call_args_list[1].args[2]
    assert "ระบบอ่านคำตอบก่อนหน้าของคุณไม่สำเร็จ" in repair
    assert "ขอโทษครับ ผมไม่เข้าใจ" in repair, "the model must see its own broken reply"
    assert "u" in repair, "and the original question, or it answers a new one"


def test_json_that_parses_into_a_hole_is_not_an_answer(client):
    """A required key that came back missing is the failure that reaches the CEO
    looking like an answer — a blank section he reads as 'nothing to report'."""
    from app import llm
    replies = [_chat_reply('{"options": []}'),
               _chat_reply('{"options": ["ก"], "recommended": "ก"}')]
    with patch("app.llm.chat", side_effect=replies) as m:
        out = llm.chat_json("anthropic", "s", "u", required=("options", "recommended"))
    assert out["data"] == {"options": ["ก"], "recommended": "ก"} and m.call_count == 2
    assert "options" in m.call_args_list[1].args[2]      # names the keys it wants


def test_a_cut_reply_is_re_asked_on_a_bigger_budget(client):
    from app import llm
    replies = [_chat_reply('{"a": 1, "b":', truncated=True),
               _chat_reply('{"a": 1, "b": 2}')]
    with patch("app.llm.chat", side_effect=replies) as m:
        out = llm.chat_json("anthropic", "s", "u", required=("a", "b"), max_tokens=4096)
    assert out["data"] == {"a": 1, "b": 2}
    assert m.call_args_list[1].kwargs["max_tokens"] == 8192


def test_a_dead_provider_is_not_asked_to_reformat(client):
    """A missing key or a 401 is not a formatting problem; re-asking just makes
    the CEO wait for the same error twice."""
    from app import llm
    with patch("app.llm.chat", return_value=_chat_reply("⚠️ ล้มเหลว", ok=False)) as m:
        out = llm.chat_json("anthropic", "s", "u", required=("a",))
    assert out["data"] is None and m.call_count == 1


def test_the_best_read_survives_a_failed_repair(client):
    """If the second try is worse than the first, keep the first — a partial
    answer beats none, as long as it is a partial answer somebody wrote."""
    from app import llm
    replies = [_chat_reply('{"a": 1}'), _chat_reply("ยังไม่เข้าใจอีก")]
    with patch("app.llm.chat", side_effect=replies):
        out = llm.chat_json("anthropic", "s", "u", required=("a", "b"))
    assert out["data"] == {"a": 1} and out["error"]


# ── the same fixes, seen from the boardroom ──

def test_a_chatty_moderator_no_longer_convenes_the_whole_board(client):
    """Stage 1 failing open calls every seat — which is the *expensive* failure:
    five models debating a question two of them have no lane for. A moderator that
    wrapped its JSON in a sentence used to trigger it."""
    chatty = "ได้ครับ นี่คือการตั้งกรอบ:\n```json\n" + _frame_json(["cfo", "coo"]) + "\n```\nหวังว่าจะช่วยได้"
    s = _start(client)
    with patch("app.llm.chat", return_value=_reply(chatty)) as m:
        s = _advance(client, s["id"])
    assert m.call_count == 1, "a readable reply must not cost a repair round"
    assert set(s["steps"][0]["results"]["framer"]["seats"]) == {"cfo", "coo"}


def test_seats_named_in_prose_still_convene(client):
    """"cfo, coo" and ["cfo","coo"] mean the same thing to everyone except a
    list comprehension over a string, which convenes one board per letter."""
    framing = json.dumps({**json.loads(_frame_json()), "seats": "cfo, coo"},
                         ensure_ascii=False)
    s = _start(client)
    with patch("app.llm.chat", return_value=_reply(framing)):
        s = _advance(client, s["id"])
    assert set(s["steps"][0]["results"]["framer"]["seats"]) == {"cfo", "coo"}


def test_a_memory_id_quoted_as_a_string_is_still_a_citation(client):
    """`"1" in {1}` is False, so every conflict a model quoted as a string was
    thrown out as a hallucination and the CEO was told the archive was clean."""
    from app import memory, store
    store.add_memory({"consult_id": 1, "question": "q", "project": None,
                      "conclusion": "c", "stance": "ทำ", "confidence": 50,
                      "constraints": [], "open_questions": [], "tripwires": []})
    quoted = json.dumps({"conflicts": [{"memory_id": "1", "past": "x", "tension": "y"}],
                         "carry_forward": []}, ensure_ascii=False)
    with patch("app.llm.chat", return_value=_reply(quoted)):
        out = memory.conflicts("คำถามใหม่", None, "anthropic")
    assert [c["memory_id"] for c in out["conflicts"]] == [1]


def test_a_clean_archive_does_not_buy_a_repair_round(client):
    """An empty conflicts list is the correct answer most of the time. Demanding a
    non-empty one would pressure the model into inventing a clash."""
    from app import memory, store
    store.add_memory({"question": "q", "project": None, "conclusion": "c"})
    with patch("app.llm.chat",
               return_value=_reply('{"conflicts": [], "carry_forward": []}')) as m:
        out = memory.conflicts("คำถามใหม่", None, "anthropic")
    assert out["conflicts"] == [] and m.call_count == 1


@pytest.mark.parametrize(("text", "expected"), [
    ("ความมั่นใจ: **62%**", 62),              # markdown bold between label and score
    ("ความมั่นใจต่อจุดยืนนี้ ประมาณ 62%", 62),  # words between them
    ("Confidence: 7/10", 70),                 # a score, not seven percent
    ("ความมั่นใจ 85/100", 85),
    ("ความมั่นใจ\n70%", 70),                   # on the next line
    ("ความมั่นใจ\n\nเหตุผล ราคา 45 บาท", None),  # not a number from further down
])
def test_confidence_survives_the_way_seats_actually_write_it(client, text, expected):
    """An unrated seat drops out of the board average, so a regex that only
    matched `label: digits` computed confidence from whichever seats happened to
    punctuate the way it expected."""
    from app import depts
    assert depts.parse_confidence(text) == expected


# ── Pipeline: recurring work, one work tree per routine ──

TRACE_JSON = json.dumps({
    "understanding": "สรุปยอดขายรายสัปดาห์แยกตามช่องทางให้ CEO อ่านเช้าวันจันทร์",
    "steps": [{"step": "ดึงยอดขาย 4 สัปดาห์", "why": "ต้องมีฐานเทียบ", "found": "โต 8%"},
              {"step": "แยกตามช่องทาง", "why": "งบโฆษณาผูกกับช่องทาง", "found": "ออนไลน์ 62%"}],
    "assumptions": ["ยอดสาขาใหม่ใกล้เคียงสาขาเดิม"],
    "evidence_used": ["เอกสาร: ยอดขาย ก.ค."],
    "unknowns": ["ยอดคืนสินค้ายังไม่เข้าระบบ"],
    "answer": "ยอดขายสัปดาห์นี้ 120,000 บาท — ออนไลน์ 62% หน้าร้าน 38%",
    "next_actions": [{"action": "ขอตัวเลขคืนสินค้า", "owner": "คุณหนึ่ง", "due": "ศุกร์นี้"}],
    "confidence": 72,
    "self_check": "ถ้าผิด จะผิดที่สมมติฐานยอดสาขาใหม่ก่อน",
    "changed_from_last": "",
}, ensure_ascii=False)


def _trace_reply(**over):
    body = json.loads(TRACE_JSON)
    body.update(over)
    return _reply(json.dumps(body, ensure_ascii=False))


def _routine(client, **over):
    payload = {"name": "รายงานยอดขายรายสัปดาห์", "owner": "คุณหนึ่ง", "dept": "cmo",
               "project": None, "goal": "รู้ยอดจริงก่อนประชุมเช้าวันจันทร์",
               "cadence": "ทุกวันจันทร์"}
    payload.update(over)
    r = client.post("/api/pipeline/routines", json=payload)
    assert r.status_code == 200, r.text
    return r.json()


def _task(client, rid, title="สรุปยอดสัปดาห์นี้", brief="แยกตามช่องทาง"):
    r = client.post(f"/api/pipeline/routines/{rid}/tasks",
                    json={"title": title, "brief": brief})
    assert r.status_code == 200, r.text
    return r.json()["tasks"][-1]


def test_a_routine_is_its_own_work_tree(client):
    """Two routines share nothing — not tasks, not runs, not history. A lane the
    CEO deletes must not take another lane's work with it."""
    a = _routine(client, name="Weekly sales")
    b = _routine(client, name="Monthly cash", owner="คุณสอง", dept="cfo")
    assert a["tree"] != b["tree"]
    assert a["tree"].startswith("routine/") and str(a["id"]) in a["tree"]

    _task(client, a["id"], "งานของ A")
    _task(client, b["id"], "งานของ B")
    client.delete(f"/api/pipeline/routines/{a['id']}")

    left = client.get("/api/pipeline").json()["routines"]
    assert [r["id"] for r in left] == [b["id"]]
    assert [t["title"] for t in left[0]["tasks"]] == ["งานของ B"]


def test_a_thai_named_tree_is_not_given_an_invented_english_handle(client):
    """`routine/3-` and `routine/3-tree` are both worse than `routine/3`."""
    r = _routine(client, name="รายงานยอดขาย")
    assert r["tree"] == f"routine/{r['id']}"


def test_a_routine_needs_a_project_and_a_named_owner(client):
    """Work with no owner is a wish. The API refuses rather than filing it."""
    assert client.post("/api/pipeline/routines",
                       json={"name": "x", "owner": "  ", "dept": "cmo"}).status_code == 400
    assert client.post("/api/pipeline/routines",
                       json={"name": " ", "owner": "y", "dept": "cmo"}).status_code == 400
    assert client.post("/api/pipeline/routines",
                       json={"name": "x", "owner": "y", "dept": "nope"}).status_code == 400

    r = _routine(client, project="YourFin")
    assert r["project"] == "YourFin" and r["owner"] == "คุณหนึ่ง"
    # the seat that will do the work, and the model behind it, travel with it
    assert r["seat_name"] == config.DEPTS["cmo"]["name"]
    assert r["provider"] == client.get("/api/state").json()["depts"][0]["provider"]


def test_a_task_inherits_the_routines_owner_rather_than_going_unassigned(client):
    r = _routine(client)
    t = _task(client, r["id"])
    assert t["owner"] == "คุณหนึ่ง" and t["status"] == "todo"
    named = client.post(f"/api/pipeline/routines/{r['id']}/tasks",
                        json={"title": "งานที่มอบต่อ", "owner": "คุณสาม"}).json()
    assert named["tasks"][-1]["owner"] == "คุณสาม"


def test_a_run_carries_the_reasoning_not_just_the_answer(client):
    """The whole point of the page: a conclusion the CEO cannot inspect is a
    conclusion he has to take on faith."""
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_trace_reply()):
        out = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run",
                          json={}).json()
    run = out["run"]
    assert run["ok"] and run["n"] == 1 and run["trigger"] == "สั่งรันเอง"
    trace = run["trace"]
    assert trace["answer"].startswith("ยอดขายสัปดาห์นี้")
    assert trace["confidence"] == 72
    assert [s["step"] for s in trace["steps"]][0] == "ดึงยอดขาย 4 สัปดาห์"
    assert trace["assumptions"] and trace["unknowns"] and trace["self_check"]
    assert trace["next_actions"][0]["owner"] == "คุณหนึ่ง"
    # …and a manifest of what it was actually allowed to read
    kinds = [i["kind"] for i in run["inputs"]]
    assert "routine" in kinds and "task" in kinds
    assert run["prompt_chars"] > 0
    # a finished run parks the task in front of the CEO, not in "done"
    assert out["routine"]["tasks"][0]["status"] == "review"


def test_the_seat_is_told_who_owns_the_work_and_what_it_is_for(client):
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_trace_reply()) as m:
        client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run", json={})
    prompt = m.call_args.args[2]
    for expected in ("คุณหนึ่ง", "รายงานยอดขายรายสัปดาห์", "ทุกวันจันทร์",
                     "รู้ยอดจริงก่อนประชุมเช้าวันจันทร์", "สรุปยอดสัปดาห์นี้"):
        assert expected in prompt, expected


def test_a_comment_reruns_the_task_and_keeps_both_answers(client):
    """A correction the CEO can no longer compare against is a correction he
    cannot check. Run 1 stays next to run 2, and the model has to say what it
    changed."""
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_trace_reply()):
        client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run", json={})

    fixed = _trace_reply(answer="ยอดขาย 120,000 บาท แยกช่องทางแล้ว",
                         changed_from_last="เพิ่มการแยกช่องทางตามที่ CEO สั่ง")
    with patch("app.llm.chat", return_value=fixed) as m:
        out = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/comment",
                          json={"text": "ยังไม่แยกตามช่องทาง แก้ใหม่"}).json()

    prompt = m.call_args.args[2]
    assert "ยังไม่แยกตามช่องทาง แก้ใหม่" in prompt, "the comment must reach the model"
    assert "120,000" in prompt, "…alongside the answer it is rejecting"

    task = out["routine"]["tasks"][0]
    assert [x["n"] for x in task["runs"]] == [1, 2], "run 1 must survive its correction"
    assert task["runs"][1]["trigger"].startswith("คอมเมนต์ของ CEO")
    assert task["runs"][1]["trace"]["changed_from_last"]
    assert task["comments"][0]["answered_by"] == 2
    assert task["open_comments"] == []


def test_a_comment_left_unrun_stays_visibly_unanswered(client):
    """`rerun: false` files the correction without acting on it — it must show as
    outstanding, not vanish into a thread."""
    r = _routine(client)
    t = _task(client, r["id"])
    out = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/comment",
                      json={"text": "ยังไม่ตรง", "rerun": False}).json()
    assert out["run"] is None
    task = out["routine"]["tasks"][0]
    assert len(task["open_comments"]) == 1 and task["comments"][0]["answered_by"] is None
    assert client.get("/api/pipeline").json()["stats"]["open_comments"] == 1

    with patch("app.llm.chat", return_value=_trace_reply()) as m:
        client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run", json={})
    assert "ยังไม่ตรง" in m.call_args.args[2], "a pending correction must reach the next run"
    assert client.get("/api/pipeline").json()["stats"]["open_comments"] == 0


def test_an_empty_comment_is_refused(client):
    r = _routine(client)
    t = _task(client, r["id"])
    assert client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/comment",
                       json={"text": "   "}).status_code == 400


def test_a_task_branches_so_two_answers_can_be_held_at_once(client):
    """Same as the Boardroom's branch, one level down: try another angle without
    destroying the answer you already have."""
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_trace_reply()):
        client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run", json={})
        client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run", json={})

    out = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/branch",
                      json={"run": 1, "title": "ลองอีกทาง"}).json()
    original, branch = out["routine"]["tasks"]
    assert [x["n"] for x in original["runs"]] == [1, 2], "the original keeps everything"
    assert [x["n"] for x in branch["runs"]] == [1], "the branch starts where it forked"
    assert branch["parent_task"] == t["id"] and branch["branched_from"] == 1
    assert branch["owner"] == original["owner"]
    # a run that never happened is not a fork point
    assert client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/branch",
                       json={"run": 9}).status_code == 400


def test_a_failed_run_is_recorded_as_a_run(client):
    """A task that silently stays 'todo' after the CEO pressed the button is the
    worst of both worlds — nothing happened and nothing says so."""
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_reply("ขอโทษครับ ผมไม่เข้าใจ")):
        out = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run",
                          json={}).json()
    assert out["run"]["ok"] is False and out["run"]["error"]
    task = out["routine"]["tasks"][0]
    assert task["status"] == "blocked" and len(task["runs"]) == 1


def test_a_messy_trace_is_coerced_rather_than_discarded(client):
    """A model that bullets its steps as prose, or writes confidence as "72%",
    has still done the reasoning — dropping the trace over punctuation hides the
    thing this page exists to show."""
    messy = _reply(json.dumps({
        "understanding": "เข้าใจแล้ว",
        "steps": "ดึงยอด\nแยกช่องทาง",
        "assumptions": "ยอดสาขาใหม่ใกล้เคียงเดิม",
        "evidence_used": [],
        "answer": "ยอด 120,000",
        "next_actions": "ขอตัวเลขคืนสินค้า",
        "confidence": "72%",
    }, ensure_ascii=False))
    r = _routine(client)
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=messy):
        run = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run",
                          json={}).json()["run"]
    trace = run["trace"]
    assert run["ok"]
    assert [s["step"] for s in trace["steps"]] == ["ดึงยอด", "แยกช่องทาง"]
    assert trace["assumptions"] == ["ยอดสาขาใหม่ใกล้เคียงเดิม"]
    assert trace["next_actions"] == [{"action": "ขอตัวเลขคืนสินค้า", "owner": "", "due": ""}]
    assert trace["confidence"] == 72


def test_a_run_reads_the_project_library_and_the_boards_past_rulings(client):
    """The routine's project is what scopes it: a lane about YourFin must not be
    fed FlowerVending paperwork, and must respect what the board already ruled."""
    from app import docs, store
    body = "ยอดขาย YourFin เดือน ก.ค. 480,000 บาท"
    docs.save_document("sales-july.txt", body.encode(), "text/plain",
                       source="upload", text=body, project="YourFin")
    store.add_memory({"consult_id": 1, "question": "ควรขยายสาขาไหม", "project": "YourFin",
                      "conclusion": "ชะลอการขยายจนกว่าจะมีตัวเลข 6 เดือน",
                      "stance": "ไม่ทำ", "confidence": 80, "constraints": [],
                      "open_questions": [], "tripwires": []})
    r = _routine(client, project="YourFin")
    t = _task(client, r["id"])
    with patch("app.llm.chat", return_value=_trace_reply()) as m:
        run = client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/run",
                          json={}).json()["run"]
    prompt = m.call_args.args[2]
    assert "480,000" in prompt and "ชะลอการขยาย" in prompt
    kinds = [i["kind"] for i in run["inputs"]]
    assert "docs" in kinds and "memory" in kinds, \
        "the manifest is what makes 'the AI ignored my document' checkable"


def test_pipeline_bad_inputs(client):
    r = _routine(client)
    t = _task(client, r["id"])
    assert client.post("/api/pipeline/routines/999/tasks", json={"title": "x"}).status_code == 404
    assert client.post(f"/api/pipeline/routines/{r['id']}/tasks",
                       json={"title": "  "}).status_code == 400
    assert client.post(f"/api/pipeline/routines/{r['id']}/tasks/99/run",
                       json={}).status_code == 404
    assert client.patch(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}",
                        json={"status": "zzz"}).status_code == 400
    assert client.patch(f"/api/pipeline/routines/{r['id']}",
                        json={"owner": "  "}).status_code == 400
    assert client.patch(f"/api/pipeline/routines/{r['id']}",
                        json={"status": "zzz"}).status_code == 400
    assert client.delete("/api/pipeline/routines/999").status_code == 404


def test_a_paused_routine_stays_listed_but_an_archived_one_steps_aside(client):
    r = _routine(client)
    client.patch(f"/api/pipeline/routines/{r['id']}", json={"status": "paused"})
    assert client.get("/api/pipeline").json()["routines"][0]["status"] == "paused"
    client.patch(f"/api/pipeline/routines/{r['id']}", json={"status": "archived"})
    assert client.get("/api/pipeline").json()["routines"] == []
    assert len(client.get("/api/pipeline?include_archived=true").json()["routines"]) == 1


def test_pipeline_counters_reach_the_shared_state_call(client):
    """The sidebar badge for outstanding corrections is fed from /api/state, so
    it has to be there whether or not the CEO is on the Pipeline page."""
    r = _routine(client)
    t = _task(client, r["id"])
    client.post(f"/api/pipeline/routines/{r['id']}/tasks/{t['id']}/comment",
                json={"text": "แก้ด้วย", "rerun": False})
    stats = client.get("/api/state").json()["pipeline"]
    assert stats["routines"] == 1 and stats["tasks"] == 1 and stats["open_comments"] == 1
